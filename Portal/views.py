import hashlib, hmac
import qrcode as qr
import string
import random
import openpyxl #to read and write excel files
from openpyxl.utils import get_column_letter #for auto-fitting column width when write exceel
from openpyxl.styles import Alignment
from io import BytesIO
from random import randint
from time import time
from django.urls import resolve, reverse_lazy
from django.conf import settings
from django.core import serializers
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db import transaction, IntegrityError
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from datetime import datetime
from django.shortcuts import render, redirect
from django.template import RequestContext
from django.db.models import Count, Q, Max
from django.contrib.auth import authenticate, login, logout
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode
from django.utils.encoding import smart_bytes, smart_str
from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage
from django.contrib.sites.shortcuts import get_current_site
from django.views.decorators.csrf import csrf_exempt, csrf_exempt
import math
import re

from django.template.loader import render_to_string
from sequences import get_next_value
from datetime import datetime
from dateutil.relativedelta import relativedelta
from calendar import monthrange
import subprocess
from PIL import Image, ImageDraw
from qrcode.image.styles.moduledrawers.pil import CircleModuleDrawer
from qrcode.image.styledpil import StyledPilImage
from dateutil.relativedelta import relativedelta

from Portal.token import password_reset_token
from .models import *
from .forms import *
from .utils import AddressMapping, AgeCalculator, CheckMemberConditions, CheckUniquePackage, GenericLibraries, DearTimeDbConn, GetCoverages, GetMedicalPlans, GetMembers, GetCompanies, PremiumCalculator

import pandas as pd
import uuid, datetime, os, json, re, xlsxwriter, base64, logging
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.http import HttpResponse
from num2words import num2words
from stat import S_IREAD
import subprocess

logging.basicConfig
logger = logging.getLogger('viewsLogger')
from Crypto.Cipher import AES
from Crypto.Protocol.KDF import PBKDF2
from Crypto.Hash import SHA256
from Crypto.Util.Padding import unpad,pad
from Crypto.Random import get_random_bytes
from django.utils.crypto import get_random_string
from firebase_dynamic_links import DynamicLinks
from cryptography.hazmat.primitives import padding
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
import phpserialize

# Create your views here.
def handler404(request, *args, **argv):
    response = render('404.html', {'host_address': settings.HOST_ADDRESS},
                                  context_instance=RequestContext(request))
    response.status_code = 404
    return response

def handler500(request, *args, **argv):
    response = render('500.html', {'host_address': settings.HOST_ADDRESS},
                                  context_instance=RequestContext(request))
    response.status_code = 500
    return response

def handler403(request, *args, **argv):
    response = render('403.html', {'host_address': settings.HOST_ADDRESS},
                                  context_instance=RequestContext(request))
    response.status_code = 403
    return response

def handler400(request, *args, **argv):
    response = render('400.html', {'host_address': settings.HOST_ADDRESS},
                                  context_instance=RequestContext(request))
    response.status_code = 400
    return response

def handler503(request, *args, **argv):
    response = render('503.html', {'host_address': settings.HOST_ADDRESS},
                                  context_instance=RequestContext(request))
    response.status_code = 503
    return response

def handler401(request, *args, **argv):
    response = render('401.html', {'host_address': settings.HOST_ADDRESS},
                                  context_instance=RequestContext(request))
    response.status_code = 401
    return response

def RefferalView(request):
    try:
        getCompanyObj       = CorporateProfile.objects.get(user_id=request.user.id)
        deartimeDB          = DearTimeDbConn()
        isConnected         = deartimeDB.connect()
        if not isConnected:
            messages.error(request, settings.CONNECTION_LOST_MESSAGE)
            return redirect(settings.HOST_ADDRESS)
        getIndividual   = deartimeDB.exec_SQL('getIndividual', {'USER_ID': getCompanyObj.deartime_payerid}, 'fetchone')
        getBankAccount  = deartimeDB.exec_SQL('validateBankAccount', {'OWNER_ID': getIndividual['dset'][0]}, 'fetchone')
        bankExist = False
        if getBankAccount['dset']:
            bankExist = True
        referral_code = GenericLibraries.insertReferralCode(getCompanyObj, deartimeDB)
        referralLink = GenericLibraries().generateReferralCodeQR(getCompanyObj.company_name, referral_code)

        referralList = []
        getReferral = deartimeDB.exec_SQL('getReferral', {'USER_ID': getCompanyObj.deartime_payerid}, type='fetchall')
        if getReferral['dset']:
            for i in range(len(getReferral['dset'])):
                jsonData = {
                    'referee_name': getReferral['dset'][i][0],
                    'created_at'  : getReferral['dset'][i][1],
                    'thanksgiving': int(getReferral['dset'][i][2] / 10),
                }
                referralList.append(jsonData)
        referralList = Paginator(referralList,10)

        paymentList = []
        getPayment = deartimeDB.exec_SQL('getPayment', {'USER_ID': getCompanyObj.deartime_payerid}, type='fetchall')
        if getPayment['dset']:
            for i in range(len(getPayment['dset'])):
                paymentListPerMonth = []
                monthExist = False
                dataDict = {
                    'month'             : getPayment['dset'][i][0],
                    'to_referee'        : getPayment['dset'][i][5],
                    'individualAmount'  : getPayment['dset'][i][2],
                }
                # add amount of the month without insert the same month to the list
                for j in range(len(paymentList)):
                    
                    if getPayment['dset'][i][0] == paymentList[j]['month'] and getPayment['dset'][i][3] == paymentList[j]['year']:
                        paymentList[j]['amount'] += getPayment['dset'][i][2]
                        paymentList[j]['paymentListPerMonth'].append(dataDict)
                        monthExist = True
                if len(paymentList) == 0 or not monthExist:
                    paymentListPerMonth.append(dataDict)
                    jsonData = {
                        'month'              : getPayment['dset'][i][0],
                        'referenceNo'        : getPayment['dset'][i][1],
                        'amount'             : getPayment['dset'][i][2],
                        'year'               : getPayment['dset'][i][3],
                        'transactionDate'    : getPayment['dset'][i][4],
                        'paymentListPerMonth': paymentListPerMonth,
                    }
                    paymentList.append(jsonData)
        
        # get distinct year and sort year
        currentYear = datetime.date.today().year
        year = []
        for i in range(len(paymentList)):
            if paymentList[i]['year'] not in year:
                year.append(paymentList[i]['year'])
        if currentYear not in year:
            year.append(currentYear)
        year.sort()
        
        # make sure every month is in the list and it will start from December
        months = ['December','November','October','September','August','July','June','May','April','March','February','January']
        
        paymentListWithYear = []
        for k in range(len(year)):
            totalAmount = 0
            newPaymentList = []
            for i in range(len(months)):
                monthExist = False
                for j in range(len(paymentList)):
                    if months[i] == paymentList[j]['month'] and year[k] == paymentList[j]['year']:
                        monthExist = True
                        break
                if not monthExist:
                    jsonData = {
                        'month'              : months[i],
                        'referenceNo'        : None,
                        'amount'             : None,
                        'transactionDate'    : None,
                        'paymentListPerMonth': None,
                    }
                    newPaymentList.append(jsonData)
                else:
                    jsonData = {
                        'month'              : months[i],
                        'referenceNo'        : paymentList[j]['referenceNo'],
                        'amount'             : round(paymentList[j]['amount'],2),
                        'transactionDate'    : paymentList[j]['transactionDate'],
                        'paymentListPerMonth': paymentList[j]['paymentListPerMonth'],
                    }
                    totalAmount += paymentList[j]['amount']
                    newPaymentList.append(jsonData)
            jsonData = {
                'year'          : year[k],
                'paymentPerYear': newPaymentList,
                'totalAmount'  : round(totalAmount,2),
            }
            paymentListWithYear.append(jsonData)
        
        # key = '59awYuceyco6kPBJ'
        # key = key.encode('ascii')
        # iv = 'ZgEineDrwLyLfX2e'
        # iv = iv.encode('ascii')
        bank_account = None
        if getBankAccount['dset']:
            # obj2 = AES.new(key,AES.MODE_CBC,iv)
            # plaintext = base64.b64decode(getBankAccount['dset'][1])
            # plaintext = unpad(obj2.decrypt(plaintext),256)
            # plaintext = base64.b64decode(plaintext)
            # bank_account = plaintext.decode('ascii')

            app_key_base64 = "Muk2HrX3OahhiGkpT9WFIpB/DqEeSKAk2aQbJ9gCpYo="
            key = base64.b64decode(app_key_base64)
            plaintext = (getBankAccount['dset'][1])
            decoded_data = base64.b64decode(plaintext)
            json_data = json.loads(decoded_data)

            # Extract IV and encrypted value
            iv = base64.b64decode(json_data['iv'])
            encrypted_value = base64.b64decode(json_data['value'])

            # Create the cipher object and decrypt the data
            cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
            decryptor = cipher.decryptor()
            padded_data = decryptor.update(encrypted_value) + decryptor.finalize()

            # Unpad the data
            unpadder = padding.PKCS7(128).unpadder()
            decrypted = unpadder.update(padded_data) + unpadder.finalize()

            # Unserialize the decrypted data
            unserializedData = phpserialize.loads(decrypted)
            bankDet = unserializedData.decode('utf-8')

            bank_account = re.sub(r'(\d{4})(?=\d)', r'\1 ', bankDet)
            
        context = {
            'company'       : getCompanyObj,
            'host_address'  : settings.HOST_ADDRESS,
            'companySalt'   : GenericLibraries().saltEncode(getCompanyObj),
            'bankExist'     : bankExist,
            'referralLink'  : referralLink,
            'referralCode'  : referral_code,
            'referralList'  : referralList,
            'paymentList'   : paymentListWithYear,
            'years'         : year,
            'currentYear'   : currentYear,
            'accountNo'     : bank_account,
        }
        deartimeDB.close()
        return render(request,'Referral/BankAccProvided.html',context)
    except Exception as ex:
        messages.error(request, str(ex))
        return redirect(settings.HOST_ADDRESS)

def RefferalViewWithoutBank(request):
    try:
        getCompanyObj       = CorporateProfile.objects.get(user_id=request.user.id)
        deartimeDB          = DearTimeDbConn()
        isConnected         = deartimeDB.connect()
        if not isConnected:
            messages.error(request, settings.CONNECTION_LOST_MESSAGE)
            return redirect(settings.HOST_ADDRESS)
        getIndividual   = deartimeDB.exec_SQL('getIndividual', {'USER_ID': getCompanyObj.deartime_payerid}, 'fetchone')
        getBankAccount  = deartimeDB.exec_SQL('validateBankAccount', {'OWNER_ID': getIndividual['dset'][0]}, 'fetchone')
        bankExist = False
        if getBankAccount['dset']:
            bankExist = True
        
        context = {
            'company'       : getCompanyObj,
            'host_address'  : settings.HOST_ADDRESS,
            'companySalt'   : GenericLibraries().saltEncode(getCompanyObj),
            'bankExist'     : bankExist,
        }
        deartimeDB.close()
        return render(request,'Referral/BankAccNotProvided.html',context)
    except Exception as ex:
        messages.error(request, str(ex))
        return redirect(settings.HOST_ADDRESS)

def GeneralView(request):
    try:
        #key and iv for encryption and decryption
        # key = get_random_bytes(32)
        # key = '59awYuceyco6kPBJ'
        # key = key.encode('ascii')
        # iv = 'ZgEineDrwLyLfX2e'
        # iv = iv.encode('ascii')
        getCompanyObj       = CorporateProfile.objects.get(user_id=request.user.id)
        deartimeDB          = DearTimeDbConn()
        isConnected         = deartimeDB.connect()
        if not isConnected:
            messages.error(request, settings.CONNECTION_LOST_MESSAGE)
            return redirect(settings.HOST_ADDRESS)
        
        queryCity       = dict(deartimeDB.exec_SQL('cities', False, type='fetchall')['dset'])
        queryPost       = dict(deartimeDB.exec_SQL('postal_codes', False, type='fetchall')['dset'])
        rawQueryState   = deartimeDB.exec_SQL('states', False, type='fetchall')['dset']
        getCompanySCP   = [getCompanyObj.state, getCompanyObj.city, getCompanyObj.postcode]
        queryState      = []
        for data in rawQueryState:
            for i in data:
                queryState.append(i)
                
        getCompanyObj.payment_due_date = datetime.datetime.strptime(getCompanyObj.payment_due_date,'%Y-%m-%d').date()

        months = [1,2,3,4,5,6,7,8,9,10,11,12]
        getDays = []
        today = datetime.datetime.now()
        year = today.year
        for month in months:
            day=monthrange(year,month)
            formDict = {
                'days'  : day[1]
            }
            getDays.append(formDict)
        
        getIndividual   = deartimeDB.exec_SQL('getIndividual', {'USER_ID': getCompanyObj.deartime_payerid}, 'fetchone')
        getBankAccount  = deartimeDB.exec_SQL('validateBankAccount', {'OWNER_ID': getIndividual['dset'][0]}, 'fetchone')
        #decrypt
        bankExist    = False
        bank_account = None
        bank_name    = None
        if getBankAccount['dset']:
            # obj2 = AES.new(key,AES.MODE_CBC,iv)
            # plaintext = base64.b64decode(getBankAccount['dset'][1])
            # plaintext = unpad(obj2.decrypt(plaintext),256)
            # plaintext = base64.b64decode(plaintext)
            # bank_account = plaintext.decode('ascii')
            # bank_account = re.sub(r'(\d{4})(?=\d)', r'\1 ', bank_account)

            # plaintext = base64.b64decode(getBankAccount['dset'][2])
            # plaintext = unpad(obj2.decrypt(plaintext),256)
            # plaintext = base64.b64decode(plaintext)
            # bank_name = plaintext.decode('ascii')
            # print(bank_name)

            bankDetails = []
            for i in range(1, 3):
                app_key_base64 = "Muk2HrX3OahhiGkpT9WFIpB/DqEeSKAk2aQbJ9gCpYo="
                key = base64.b64decode(app_key_base64)
                plaintext = (getBankAccount['dset'][i])
                decoded_data = base64.b64decode(plaintext)
                json_data = json.loads(decoded_data)

                # Extract IV and encrypted value
                iv = base64.b64decode(json_data['iv'])
                encrypted_value = base64.b64decode(json_data['value'])

                # Create the cipher object and decrypt the data
                cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
                decryptor = cipher.decryptor()
                padded_data = decryptor.update(encrypted_value) + decryptor.finalize()

                # Unpad the data
                unpadder = padding.PKCS7(128).unpadder()
                decrypted = unpadder.update(padded_data) + unpadder.finalize()

                # Unserialize the decrypted data
                unserializedData = phpserialize.loads(decrypted)
                bankDet = unserializedData.decode('utf-8')
                bankDetails.append(bankDet)

            bankExist = True
        
            bank_account = bankDetails[0]
            bank_name = bankDetails[1]

            bank_account = re.sub(r'(\d{4})(?=\d)', r'\1 ', bank_account)

        banks = ['MALAYAN BANKING BERHAD','AFFIN BANK BERHAD / AFFIN ISLAMIC BANK','AL-RAJHI BANKING & INVESTMENT CORP (M) BERHAD','ALLIANCE BANK MALAYSIA BERHAD','AmBANK BERHAD','BANK ISLAM MALAYSIA',
                 'BANK KERJASAMA RAKYAT MALAYSIA BERHAD','BANK MUAMALAT','BANK OF AMERICA','BANK OF CHINA (MALAYSIA) BERHAD','BANK PERTANIAN MALAYSIA BERHAD (AGROBANK)','BANK SIMPANAN NASIONAL BERHAD',
                 'BNP PARIBAS MALAYSIA','Bangkok Bank Berhad','CHINA CONST BK (M) BHD','CIMB BANK BERHAD','CITIBANK BERHAD','DEUTSCHE BANK (MSIA) BERHAD','HONG LEONG BANK','HSBC BANK MALAYSIA BERHAD',
                 'INDUSTRIAL & COMMERCIAL BANK OF CHINA','J.P. MORGAN CHASE BANK BERHAD','KUWAIT FINANCE HOUSE (MALAYSIA) BHD','MBSB BANK BERHAD','MIZUHO CORPORATE BANK MALAYSIA','MUFG BANK (MALAYSIA) BHD',
                 'OCBC BANK(MALAYSIA) BHD','PUBLIC BANK','RHB BANK','STANDARD CHARTERED BANK','SUMITOMO MITSUI BANKING CORPORATION MALAYSIA BHD','UNITED OVERSEAS BANK BERHAD']

        context = {
            'company'       : getCompanyObj,
            'host_address'  : settings.HOST_ADDRESS,
            'companySalt'   : GenericLibraries().saltEncode(getCompanyObj),
            'day_in_a_month': getDays,
            'cities'        : queryCity,
            'postal_codes'  : queryPost,
            'states'        : queryState,
            'bank_account'  : bank_account,
            'bank_name'     : bank_name,
            'scp'           : getCompanySCP,
            'bankExist'     : bankExist,
            'banks'         : banks,
        }

        if request.method == 'POST':
            if request.POST.get('modified_contact'):
                country_code = "+60"
                concat_mobile_number = request.POST['modified_countryCode'] + request.POST['modified_contact']
                formatted_mobile_number = re.sub('[^0-9+]+', '', str(concat_mobile_number))    
                if country_code in formatted_mobile_number:
                    formatted_mobile_number = formatted_mobile_number[2:]
                # getExistEmailOrMobile = deartimeDB.exec_SQL('validateExistingEmailMobile', {'MOBILE': formatted_mobile_number, 'EMAIL' : 'email'}, 'fetchone')
                getAllCorporateProfile = CorporateProfile.objects.all().filter(rejected=False)
                for company in getAllCorporateProfile:
                    if company.contact1 == formatted_mobile_number:
                        messages.error(request, 'Phone number exist!')
                        return render(request,'Referral/General.html',context)
                        
                getCompanyObj.contact1 = formatted_mobile_number
                getCompanyObj.save()
                messages.success(request, 'Update Successfully!')
                
            elif request.POST.get('address1'):
                getCompanyObj.address_line1 = request.POST['address1']
                getCompanyObj.address_line2 = request.POST['address2']
                getCompanyObj.address_line3 = request.POST['address3']
                getCompanyObj.state = request.POST['state']
                getCompanyObj.city = request.POST['city']
                getCompanyObj.postcode = request.POST['postcode']
                getCompanyObj.save()
                messages.success(request, 'Update Successfully!')

            elif (request.POST.get('companyPaymentMode') and request.POST.get('paymentDueDate_post')):
                payment_mode = request.POST['companyPaymentMode']
                # save the old value before get the new value
                old_payment_mode = getCompanyObj.payment_mode
                if getCompanyObj.payment_mode is None or getCompanyObj.payment_mode.strip() == '':
                    old_payment_mode = 'Yearly'
                    getCompanyObj.payment_mode = old_payment_mode
                    getCompanyObj.save()
                old_payment_due_date = getCompanyObj.payment_due_date

                current_payment_mode = getCompanyObj.payment_mode
                payment_due_date = request.POST['paymentDueDate_post']
                payment_due_date = datetime.datetime.strptime(payment_due_date, "%Y-%m-%d").date()

                # save old & new payment mode and due date into history
                getChangePaymentMode = PaymentModeHistory.objects.filter(corporate_id=getCompanyObj.id, is_void=False)

                if getChangePaymentMode:
                    for pm in getChangePaymentMode:
                        pm.is_void = True
                        pm.save()

                payment_update_history = PaymentModeHistory(
                    corporate=getCompanyObj, 
                    old_payment_mode=old_payment_mode,
                    old_payment_due_date=old_payment_due_date,
                    new_payment_mode=payment_mode,
                    new_payment_due_date=payment_due_date
                )
                payment_update_history.save()
                
                if payment_mode == 'Yearly':
                    cpfo_payment_mode = 'annually'
                elif payment_mode == 'Monthly':
                    cpfo_payment_mode = 'monthly'
                updateNewPaymentTermDict = {'CORPORATE_ID': getCompanyObj.deartime_payerid, 'NEW_PAYMENT_TERM': cpfo_payment_mode}
                deartimeDB.exec_SQL('UpdateNewPaymentTerm', updateNewPaymentTermDict, 'update')

                if old_payment_mode == payment_mode:
                    saveMessageQueue = MessagingQueue(
                        email_address = getCompanyObj.email_address,
                        module = 'ChangePaymentDueDateView'
                    )
                else:
                    saveMessageQueue = MessagingQueue(
                        email_address = getCompanyObj.email_address,
                        module = 'ChangePaymentModeView'
                    )
                saveMessageQueue.save()

                GenericLibraries().checkChangePaymentMode(getCompanyObj)

                if payment_update_history.new_payment_mode == 'Yearly' and payment_update_history.old_payment_mode == 'Monthly':
                    GenericLibraries().generateInvoiceMonthlyPayment(getCompanyObj)
                
                messages.success(request, 'Changes successfully saved!')

            elif request.POST.get('bankAccount_post') and request.POST.get('bankName_post'):
                encBankList = []
                bankAccount = request.POST['bankAccount_post']
                bankAccount = re.sub(r'[\D\s]', '', bankAccount)
                # cleaned_bank_account = re.sub(r'(\d{4})(?=\d)', r'\1 ', bankAccount)
                encBankList.append(bankAccount)
                # bankAccountBase64 = bankAccount.encode("ascii")
                # bankAccountBase64 = base64.b64encode(bankAccountBase64)
                
                #encrypt then change ciphertext to base 64 format string
                # obj = AES.new(key,AES.MODE_CBC,iv)
                # ciphertext = obj.encrypt(pad(bankAccountBase64,256))
                # ciphertextBase64 = base64.b64encode(ciphertext)
                # ciphertextBase64String = ciphertextBase64.decode('ascii')

                bankName = request.POST['bankName_post']
                encBankList.append(bankName)
                # bankNameBase64 = bankName.encode("ascii")
                # bankNameBase64 = base64.b64encode(bankNameBase64)

                # bankNameciphertext = obj.encrypt(pad(bankNameBase64,256))
                # bankNameciphertextBase64 = base64.b64encode(bankNameciphertext)
                # bankNameciphertextBase64String = bankNameciphertextBase64.decode('ascii')

                encEdBankList = []
                for deta in encBankList:
                    app_key_base64 = "Muk2HrX3OahhiGkpT9WFIpB/DqEeSKAk2aQbJ9gCpYo="
                    key = base64.b64decode(app_key_base64)

                    serializedData = phpserialize.dumps(deta)

                    numPadder = padding.PKCS7(128).padder()
                    numPaddedData = numPadder.update(serializedData) + numPadder.finalize()

                    iv = os.urandom(16)

                    numCipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
                    numEncryptor = numCipher.encryptor()
                    numEncrypted = numEncryptor.update(numPaddedData) + numEncryptor.finalize()

                    # Encode the IV and encrypted data in Base64
                    iv_base64 = base64.b64encode(iv).decode('utf-8')
                    numEncrypted_base64 = base64.b64encode(numEncrypted).decode('utf-8')

                    # Laravel uses HMAC with SHA-256 to generate the MAC
                    mac = hmac.new(key, iv_base64.encode('utf-8') + numEncrypted_base64.encode('utf-8'), hashlib.sha256).hexdigest()

                    # Combine everything into a JSON structure and Base64 encode it
                    payload = json.dumps({'iv': iv_base64, 'value': numEncrypted_base64, 'mac': mac})
                    encodedPayload = base64.b64encode(payload.encode('utf-8')).decode('utf-8')
                    encEdBankList.append(encodedPayload)

                bankNumEncoded_payload = encEdBankList[0]
                bankNameEncoded_payload = encEdBankList[1]
                
                getIndividual   = deartimeDB.exec_SQL('getIndividual', {'USER_ID': getCompanyObj.deartime_payerid}, 'fetchone')
                checkOwnerBank       = deartimeDB.exec_SQL('validateBankAccount', {'OWNER_ID': getIndividual['dset'][0]}, 'fetchone')
                if(checkOwnerBank['dset']):
                    dataDict = {'ACC_NO':bankNumEncoded_payload ,'BANK_NAME':bankNameEncoded_payload,'UPDATE_DATETIME':datetime.datetime.now(),'OWNER_ID': getIndividual['dset'][0]}
                    updateBankAccount = deartimeDB.exec_SQL('updateBankAccount',dataDict, 'update')
                    if 'error' in updateBankAccount:
                        messages.error(request, updateBankAccount['error'])
                        return render(request,'Referral/General.html',context)
                else:
                    dataDict = (str(uuid.uuid4()),getIndividual['dset'][0],'App\Individual',bankNumEncoded_payload,bankNameEncoded_payload,None,None,str(datetime.datetime.now()),str(datetime.datetime.now()),None)
                    insertBankAccount = deartimeDB.exec_SQL('insertBankAccount',dataDict,'insert')
                    if 'error' in insertBankAccount:
                        messages.error(request, insertBankAccount['error'])
                        return render(request,'Referral/General.html',context)
                bankExist = True
                context = {
                    'company'       : getCompanyObj,
                    'host_address'  : settings.HOST_ADDRESS,
                    'companySalt'   : GenericLibraries().saltEncode(getCompanyObj),
                    'day_in_a_month': getDays,
                    'cities'        : queryCity,
                    'postal_codes'  : queryPost,
                    'states'        : queryState,
                    'bank_account'  : bankAccount,
                    'bank_name'     : bankName,
                    'scp'           : getCompanySCP,
                    'bankExist'     : bankExist,
                    'banks'         : banks,
                }
                messages.success(request, 'Update Successfully!')
                return redirect(settings.HOST_ADDRESS+'/referral')
            else:
                messages.error(request, "Please fill out required field.")
            deartimeDB.close()
            return render(request,'Referral/General.html',context)
        deartimeDB.close()
        return render(request,'Referral/General.html',context)
    except Exception as ex:
        messages.error(request, str(ex))
        logger.error(str(ex),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/member-list')

@csrf_exempt
def LoginView(request):
    if 'next' in request.GET:
        messages.error(request, "Login is required.")
    else:
        if request.method == 'POST':
            user = authenticate(request, username=request.POST['email'], password=request.POST['password'])
            if user:
                if user.is_superuser == True and user.is_active == True:
                    login(request, user)
                    if 'next' in request.GET:
                        return redirect(settings.HOST_ADDRESS+'/'+request.GET['next'])
                    else:
                        return redirect(settings.HOST_ADDRESS+'/company-approval')
                elif user.is_active == True:
                    company_info      = CorporateProfile.objects.get(user_id=user.id)
                    isCompanyVerified = company_info.verified
                    isSubmitted       = company_info.submitted
                    isRejected        = company_info.rejected
                    isDeferred        = company_info.remarks
                    if isRejected:
                        messages.error(request, 'Invalid username or password!')
                        return redirect(settings.HOST_ADDRESS)
                    elif isCompanyVerified and isSubmitted:
                        login(request, user)
                        if 'next' in request.GET:
                            return redirect(settings.HOST_ADDRESS+'/'+request.GET['next'])
                        else:
                            return redirect(settings.HOST_ADDRESS+'/dashboard')
                    elif (not isCompanyVerified) and isSubmitted:
                        if isDeferred:
                            login(request, user)
                            context = {
                                'company_info' : company_info,
                                'host_address' : settings.HOST_ADDRESS,
                                'companySalt'  : GenericLibraries().saltEncode(company_info)
                            }
                            return render(request, 'CompanyRegistration/CompanyAccountDeferred.html', context)
                        else:
                            login(request, user)
                            context = {
                                'company_info' : company_info,
                                'host_address' : settings.HOST_ADDRESS,
                                'companySalt'  : GenericLibraries().saltEncode(company_info)
                            }
                            return render(request, 'CompanyRegistration/CompanyAccountSubmitted.html', context)
                    elif (not isCompanyVerified) and (not isSubmitted):
                        login(request, user)
                        companySalt = GenericLibraries().saltEncode(company_info)
                        return redirect(settings.HOST_ADDRESS+'/company-registration-login/{}'.format(companySalt))
                else:
                    getCorporateUserObj = CorporateUser.objects.get(email=request.POST['email'])
                    checkLoginAttempt = LoginAttemptControl.objects.filter(user=getCorporateUserObj)
                    if checkLoginAttempt:
                        getLoginAttemptObj = LoginAttemptControl.objects.get(user=getCorporateUserObj)
                        if getLoginAttemptObj.attempts == settings.LIMITED_LOGIN_ATTEMPT:
                            messages.error(request, 'Your account has been locked due to multiple failed login attempts. Please reset your password.')
                        else:
                            messages.error(request, 'User account status is inactive! Please contact administrator')
                    else:
                        messages.error(request, 'User account status is inactive! Please contact administrator')
                    return redirect(settings.HOST_ADDRESS)

            else:
                checkCorporateUserObj = CorporateUser.objects.filter(email=request.POST['email'])
                if checkCorporateUserObj:
                    getCorporateUserObj = CorporateUser.objects.get(email=request.POST['email'])
                    checkLoginAttempt = LoginAttemptControl.objects.filter(user=getCorporateUserObj)
                    if checkLoginAttempt:
                        getLoginAttemptObj = LoginAttemptControl.objects.get(user=getCorporateUserObj)
                        if getLoginAttemptObj.attempts == settings.LIMITED_LOGIN_ATTEMPT:
                            getCorporateUserObj.is_active = False
                            getCorporateUserObj.save()
                            messages.error(request, 'Your account has been locked due to multiple failed login attempts. Please reset your password.')
                            return redirect(settings.HOST_ADDRESS)
                        else:
                            getLoginAttemptObj.attempts = getLoginAttemptObj.attempts + 1 
                            getLoginAttemptObj.attempt_datetime = datetime.datetime.now()
                            getLoginAttemptObj.save()
                    else:
                        saveLoginAttempt = LoginAttemptControl(
                            user = getCorporateUserObj
                        )
                        saveLoginAttempt.save()
                messages.error(request, 'Invalid username or password!')
                return redirect(settings.HOST_ADDRESS)
    context = {
        'host_address': settings.HOST_ADDRESS
    }
    return render(request, 'Auth/Login.html', context)

@csrf_exempt
def ResetPasswordEmailView(request, uidb64, token):
    try:
        uid = smart_str(urlsafe_base64_decode(uidb64))
        user = CorporateUser.objects.get(id=uid)

    except (TypeError, ValueError, OverflowError, CorporateUser.DoesNotExist):
        user = None

    if user is not None and password_reset_token.check_token(user, token, '1'):
        context = {
            'email' : user.email,
            'host_address' : settings.HOST_ADDRESS
        }
        return render(request, 'Auth/SetNewPassword.html', context)
    
    else:
        if user is not None and password_reset_token.check_token(user, token, '2'):
            current_site = get_current_site(request)
            context = {
                'domain': current_site.domain,
                'scheme': request.scheme,
                'uidb64': urlsafe_base64_encode(smart_bytes(user.id)),
                'token' : password_reset_token.make_token(user),
                'flag'  : 1,
                'host_address' : settings.HOST_ADDRESS
            }
            return render(request, 'EmailLink/ResendPasswordResetLink.html', context)
        else:
            context = {
                'host_address': settings.HOST_ADDRESS
            }
            return render(request, 'EmailLink/EmailLinkExpired.html', context)

@csrf_exempt
def ResetPasswordView(request):
    if request.method == 'POST':
        email_address = request.POST.get('email_address', '')  # Get the email address from the POST data
        messages.success(request, "E-mail has been sent to " + email_address)# Show the messages no matter email sent or not to increase security
        
        if CorporateProfile.objects.filter(email_address=email_address).exists():# Check if the email address exists in the 'CorporateProfile' table
            saveMsgQueueObj = GenericLibraries.saveMessageQueue(email_address, 'ResetPasswordEmailView')
    
    context = {
        'host_address': settings.HOST_ADDRESS
    }
    return render(request, 'Auth/ResetPassword.html', context)
@csrf_exempt
def SetNewPassword(request):
    if request.method == 'POST':
        try:
            getCorporateUser = CorporateUser.objects.get(email=request.POST['email'])

            context = {
                'email' : request.POST['email'],
                'host_address' : settings.HOST_ADDRESS
            }

            if getCorporateUser.check_password(request.POST['password']):
                messages.error(request, "New password cannot be same as current password!")
                return render(request, 'Auth/SetNewPassword.html', context)
            else:
                getCorporateUser.set_password(request.POST['password'])
                getCorporateUser.is_active = True
                getCorporateUser.save()
                checkLoginAttempt = LoginAttemptControl.objects.filter(user=getCorporateUser)
                if checkLoginAttempt:
                    getLoginAttemptObj = LoginAttemptControl.objects.get(user=getCorporateUser)
                    getLoginAttemptObj.attempts = 0
                    getLoginAttemptObj.attempt_datetime = datetime.datetime.now()
                    getLoginAttemptObj.save()
                messages.success(request, "Password changed successfully!")
                return redirect(settings.HOST_ADDRESS)

        except Exception as e:
            messages.error(request, str(e))
            logger.error(str(e),extra={'username':request.user.id})
            return redirect(settings.HOST_ADDRESS+'/set-new-password')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None) 
def ChangePassword(request):
    getCorporateUser = CorporateUser.objects.get(id=request.user.id)
    context = {
                'email' : getCorporateUser.email, 
                'host_address' : settings.HOST_ADDRESS
             }
    if request.method == 'POST':
        try:
            if not getCorporateUser.check_password(request.POST['old_password']):
                messages.error(request, "The password is incorrect.")
            else:
                if request.POST['new_password'] == request.POST['old_password']:
                    messages.error(request, "New password cannot be same as current password!")
                    return render(request, 'Auth/ChangePassword.html', context)
                else:
                    getCorporateUser.set_password(request.POST['new_password'])
                    getCorporateUser.save()
                    messages.success(request, "Password changed successfully!")
                    return redirect(settings.HOST_ADDRESS)

        except Exception as e:
            messages.error(request, str(e))
            logger.error(str(e),extra={'username':request.user.id})
    return render(request, 'Auth/ChangePassword.html', context)

@csrf_exempt
def ResendLinkView(request, uidb64, token, flag):
    try:
        uid = smart_str(urlsafe_base64_decode(uidb64))
        user = CorporateUser.objects.get(id=uid)

        if flag == 1:
            saveMsgQueueObj = GenericLibraries.saveMessageQueue(user.email, 'ResetPasswordEmailView')

            context = {
                'email' : user.email,
                'flag'  : 1
            }


        elif flag == 2:
            saveMsgQueueObj = GenericLibraries.saveMessageQueue(user.email, 'SignUpValidationEmailView')

            context = {
                'email' : user.email,
                'flag'  : 2,
                'host_address' : settings.HOST_ADDRESS
            }
        
        messages.success(request, "E-mail sent successfully" if flag == 1 else "Verification e-mail sent successfully")
        return render(request, 'Auth/AwaitingEmailValidation.html', context)
    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'resend-link/{}/{}/{}'.format(uidb64, token, flag))

def cleanup(request):
    name = request.POST['companyName']
    regNum = request.POST['registrationNumber']

    excludeName = "[!@#\$%\^\*\[\]\.\|\+\?\{\}\\\/_\,\<\>`\~\- ]"
    excludeRegNum = "[!@#\$%\^&\*\(\)\[\]\.\|\+\?\{\}\\\/_\,\<\>`\~ ]"

    cleanedName = re.sub(excludeName, "", name).upper()
    cleanedRegNum = re.sub(excludeRegNum, "", regNum).upper()

    return [cleanedName, cleanedRegNum]

@csrf_exempt
def SignUpView(request):
    if request.method == 'POST':
        try:
            cleanedDetails = cleanup(request)
            email = request.POST['companyEmail']

            hasName = CorporateProfile.objects.filter(val_company_name=cleanedDetails[0])
            if not hasName:
                hasRegNum = CorporateProfile.objects.filter(val_registration_no=cleanedDetails[1])
                if not hasRegNum:    
                    hasEmail = CorporateProfile.objects.filter(email_address__iexact=email)
                    if hasEmail:
                        messages.error(request, 'Company E-mail already exist!')
                        return redirect(settings.HOST_ADDRESS+'/sign-up/')
                else:
                    messages.error(request, 'Company Registration Number already exist!')
                    return redirect(settings.HOST_ADDRESS+'/sign-up/')
            else:
                for cmp in hasName:
                    if cmp.rejected:
                        messages.error(request, 'Registration failed. Please contact DearTime administrators.')
                        return redirect(settings.HOST_ADDRESS+'/sign-up/')
                    else:
                        messages.error(request, 'Company Name already exist!')
                        return redirect(settings.HOST_ADDRESS+'/sign-up/')

            # To create user object
            saveUserObj = CorporateUser(
                email     = request.POST['companyEmail'],
                is_active = True
            )
            saveUserObj.set_password(request.POST['password'])
            saveUserObj.save()

            # Create company profile
            saveCompanyProfileObj = CorporateProfile(
                company_name    = request.POST['companyName'],
                val_company_name = cleanedDetails[0].upper(),
                registration_no = request.POST['registrationNumber'],
                val_registration_no = cleanedDetails[1].upper(),
                email_address   = request.POST['companyEmail'],
                user_id         = saveUserObj.id,
                status          = 'Incomplete Document'
            )
            saveCompanyProfileObj.save()

            saveMsgQueueObj = GenericLibraries.saveMessageQueue(request.POST['companyEmail'], 'SignUpValidationEmailView')

            context = {
                'email': request.POST['companyEmail'],
                'flag' : 2,
                'host_address' : settings.HOST_ADDRESS
            }
            return render(request, 'Auth/AwaitingEmailValidation.html', context)
        except Exception as ex:
            logger.error(str(ex),extra={'username':request.user.id})
            context = {
                'host_address': settings.HOST_ADDRESS
            }
            return render(request, 'Auth/SignUp.html', context)
    context = {
        'host_address': settings.HOST_ADDRESS
    }
    return render(request, 'Auth/SignUp.html', context)

@csrf_exempt
def CompanyRegistrationViewLogin(request, companySalt):
    try:
        decodeCompanySalt = base64.b64decode(companySalt)
        decodeASCII = decodeCompanySalt.decode('UTF-8')
        getCompanyID = decodeASCII.split('_')
        company_id = getCompanyID[1]
        getRelationship     = RelationshipType.objects.filter(is_active=True)
        getEntityType       = EntityType.objects.filter(is_active=True)
        getEnFormType       = CompanyFormType.objects.filter(is_active=True)

        entityFormData  = []
        formData        = []

        for eobj in getEntityType:
            for fobj in getEnFormType:
                if (eobj.id == fobj.entity_type_id ):
                    formData.append(fobj.form_type_id)
                    formattedForm = ','.join(formData)
                else:
                    formData = []

            formDict = {
                'entity_type_id'  : eobj.id,
                'entity_type_name': eobj.entity_name,
                'form_types'      : formattedForm,
            }
            formattedForm = ''
            entityFormData.append(formDict)

        getSQLConnection    = AddressMapping()
        isConnected         = getSQLConnection.connect()
        if not isConnected:
            messages.error(request, settings.CONNECTION_LOST_MESSAGE)
            return redirect(settings.HOST_ADDRESS)
        if request.method == 'POST':
            try:
                getCompanyInfo = GenericLibraries.registerCompany(request, company_id)
                saveRelationshipObj = GenericLibraries.saveCompanyRelationship(request, getCompanyInfo)
                saveMsgQueueObj = GenericLibraries.saveMessageQueue(getCompanyInfo.email_address, 'CompanyRegistrationView')

                for key, value in request.POST.items():
                    if (key == 'entityType_post'):
                        entity_id = int(value)

                getEntityIdInfo = EntityType.objects.get(id=entity_id)

                try:
                    # Check if directory exist
                    companyFolder    = settings.MEDIA_ROOT + '/' + getCompanyInfo.company_name + '/'
                    isDirectoryExist = os.path.exists(companyFolder)
                    if not isDirectoryExist:
                        os.mkdir(companyFolder)

                    for k, v in request.POST.items():
                        if k.startswith('base64'):
                            splitData = v.split('|')
                            filename = '/' + getCompanyInfo.company_name + '/' + splitData[0]
                            splitFile = splitData[0].split('.')
                            image_64_decode = base64.b64decode(splitData[1].split(',')[1])
                            form_code = k.split('base64-')[1].split('_0')[0]
                            open(settings.MEDIA_ROOT + filename, 'wb').write(image_64_decode)
                            new_attachment = CorporateProfileFormAttachment(
                                company         = getCompanyInfo,
                                form_code       = form_code,
                                attachment      = splitFile[0],
                                attachment_type = '.'+splitFile[1],
                            )
                            new_attachment.save()
                            if form_code not in ('brs-form', 'icPass-form', 'authorization-form'):
                                new_attachment.entity_type = getEntityIdInfo
                                new_attachment.save()

                    getCompanyInfo = GenericLibraries.updateCompanySubmittedStatus(getCompanyInfo)
                except (Exception,FileNotFoundError) as e:
                    messages.error(request, "No such file or directory.")
                    logger.error(str(e),extra={'username':request.user.id})
                    context = GenericLibraries.deleteUploadedDoc(request, getCompanyInfo, getRelationship, getSQLConnection)
                    return render(request, 'CompanyRegistration/CompanyAccountPending.html', context)
                
                return render(request, 'CompanyRegistration/CompanyAccountSubmitted.html', {'company_info' : getCompanyInfo, 'host_address' : settings.HOST_ADDRESS, 'companySalt': GenericLibraries().saltEncode(getCompanyInfo)})
            except (Exception, IntegrityError,) as e:
                messages.error(request, str(e))
                logger.error(str(e),extra={'username':request.user.id})
                # To delete any uploaded documents
                context = GenericLibraries.deleteUploadedDoc(request, getCompanyInfo, getRelationship, getSQLConnection)
                return render(request, 'CompanyRegistration/CompanyAccountPending.html', context)

        try:
            queryCity = getSQLConnection.queryCity()
            queryPost = getSQLConnection.queryPost()
            queryState = getSQLConnection.queryState()
            getCompanyInfo  = CorporateProfile.objects.get(id=company_id)
            context = {
                'path'        : request.path_info,
                'company_info': getCompanyInfo,
                'relationship': getRelationship,
                'cities'      : queryCity,
                'postal_codes': queryPost,
                'states'      : queryState,
                'entity_data' : entityFormData,
                'host_address' : settings.HOST_ADDRESS,
                'http_host_address': settings.HTTP_HOST_ADDRESS,
                'companySalt': GenericLibraries().saltEncode(getCompanyInfo)
            }
            return render(request, 'CompanyRegistration/CompanyAccountPending.html', context)
        except Exception as e:
            messages.error(request, str(e))
            logger.error(str(e),extra={'username':request.user.id})
            return redirect(settings.HOST_ADDRESS)
    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS)

@csrf_exempt
def CompanyRegistrationView(request, uidb64, token):
    try:
        uid  = smart_str(urlsafe_base64_decode(uidb64))
        user = CorporateUser.objects.get(id=uid)
    except (TypeError, ValueError, OverflowError, CorporateUser.DoesNotExist):
        user = None

    if user:
        getCompanyInfo  = CorporateProfile.objects.get(user_id=user.id)
        if getCompanyInfo.submitted == False:
            if user is not None and password_reset_token.check_token(user, token, '2'):
                companySalt = GenericLibraries().saltEncode(getCompanyInfo)
                return redirect(settings.HOST_ADDRESS+'/company-registration-login/{}'.format(companySalt))
            else:
                return render(request, 'EmailLink/EmailLinkExpired.html', {'flag': 2, 'host_address' : settings.HOST_ADDRESS})
        else:
            return render(request, 'CompanyRegistration/CompanyAccountSubmitted.html', {'company_info' : getCompanyInfo, 'host_address' : settings.HOST_ADDRESS, 'companySalt': GenericLibraries().saltEncode(getCompanyInfo)})
    else:
        context = {
            'host_address': settings.HOST_ADDRESS
        }
        return render(request, 'UserNotFound.html', context)
      
@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def UploadCorporateForm(request, companySalt, form_type):
    decodeCompanySalt = base64.b64decode(companySalt)
    decodeASCII = decodeCompanySalt.decode('UTF-8')
    getCompanyID = decodeASCII.split('_')
    company_id = getCompanyID[1]
    attachment_files = []
    with transaction.atomic():
        getCompanyObj = CorporateProfile.objects.get(id=company_id)
        for attachment in request.FILES.getlist(form_type):
            try:
                # Check if directory exist
                companyFolder    = settings.MEDIA_ROOT + '/' + getCompanyObj.company_name + '/'
                isDirectoryExist = os.path.exists(companyFolder)
                if not isDirectoryExist:
                    os.mkdir(companyFolder)

                filename = '/' + getCompanyObj.company_name + '/' + os.path.basename(attachment.name)
                open(settings.MEDIA_ROOT + filename, 'wb').write(attachment.file.read())
                split_tup = os.path.splitext(os.path.basename(attachment.name))
                new_attachment = CorporateProfileFormAttachment(
                    company         = getCompanyObj,
                    form_code       = form_type,
                    attachment      = split_tup[0],
                    attachment_type = split_tup[1],
                )
                new_attachment.save()
                attachment_files.append(settings.MEDIA_ROOT + filename)
            except (Exception,FileNotFoundError) as e:
                messages.error(request, "No such file or directory.")
                logger.error(str(e),extra={'username':request.user.id})
    return HttpResponse(json.dumps(attachment_files), content_type='application/json')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)  
def RemoveCorporateForm(request, companySalt, form_type):
    decodeCompanySalt = base64.b64decode(companySalt)
    decodeASCII = decodeCompanySalt.decode('UTF-8')
    getCompanyID = decodeASCII.split('_')
    company_id = getCompanyID[1]
    attachment_files = []
    with transaction.atomic():
        getCompanyObj = CorporateProfile.objects.get(id=company_id)
        filename = request.POST['filename']
        try:
            split_tup = os.path.splitext(os.path.basename(filename))
            removeFile = CorporateProfileFormAttachment.objects.filter(form_code=form_type, attachment=split_tup[0], company=getCompanyObj)
            removeFile.delete()
        except (Exception, FileNotFoundError) as ex:
            messages.error(request, "No such file or directory.")
            logger.error(str(ex),extra={'username':request.user.id})
    return HttpResponse(json.dumps(attachment_files), content_type='application/json')

@csrf_exempt  
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def SubmittedCompanyRegistrationView(request):
    context = {
        'host_address': settings.HOST_ADDRESS
    }
    return render(request, 'CompanyRegistration/CompanyAccountSubmitted.html', context)

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None) 
def CompanyApprovalListView(request):
    rowCount = 5
    getPendingVerificationCompany = CorporateProfile.objects.all()
    pendingCompanyLists = []
    rejectedPendingCompanyLists = []

    for cmp in getPendingVerificationCompany:
        getRelationships = CompanyRelationship.objects.filter(company_id=cmp.id)
        if len(list(getRelationships)) == 0:
            getRelationships = None
        getCompanyForms  = CorporateProfileFormAttachment.objects.filter(company_id=cmp.id)
        if len(list(getCompanyForms)) == 0:
            getCompanyForms = None
        serializerCMP    = json.loads(serializers.serialize('json', [cmp]))
        serializerCMP[0]['fields'].update({
            'relationships': getRelationships,
            'forms'        : getCompanyForms,
            'id'           : serializerCMP[0]['pk'],
            'companySalt'  : GenericLibraries().saltEncode(cmp)
        })
        
        if cmp.rejected == False:
            pendingCompanyLists.append(serializerCMP[0]['fields'])
        else:
            rejectedPendingCompanyLists.append(serializerCMP[0]['fields'])

    companyTable = Paginator(pendingCompanyLists, rowCount)
    rejectedTable = Paginator(rejectedPendingCompanyLists, rowCount)

    context = {
        'company_lists'          : companyTable,
        'rejected_company_lists' : rejectedTable,
        'path'                   : request._current_scheme_host,
        'host_address' : settings.HOST_ADDRESS
    }
    return render(request, 'CorporateApproval/CompanyApprovalList.html', context)

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def ViewCompanyAccountView(request, companySalt):
    try:
        decodeCompanySalt = base64.b64decode(companySalt)
        decodeASCII = decodeCompanySalt.decode('UTF-8')
        getCompanyID = decodeASCII.split('_')
        companyID = getCompanyID[1]
        hasAuthorizationForm = False
        hasBrdForm           = False
        hasCopdForm          = False
        has917Form           = False
        has1328Form          = False
        has2478Form          = False
        has4958Form          = False
        haslrfcForm          = False
        has8Form             = False
        hasRosForm           = False
        if request.method == 'POST':
            if request.POST['formType'] == 'Approve':
                getCompanyObj = CorporateProfile.objects.get(id=request.POST['modalCompanyID'])
                getCompanyObj.verified         = True
                getCompanyObj.updated_by       = request.user
                getCompanyObj.updated_datetime = datetime.datetime.now()
                
                # Save to Deartime DB
                deartimeDB   = DearTimeDbConn()
                isConnected  = deartimeDB.connect()
                if not isConnected:
                    messages.error(request, settings.CONNECTION_LOST_MESSAGE)
                    return redirect(settings.HOST_ADDRESS+'/company-approval') 
                getLatestID  = deartimeDB.exec_SQL('selectMaxIDUserTB', {}, 'fetchone')
                nextLatestID = getLatestID['dset'][0] + 1
                userRefNo    = 'CU' + str(nextLatestID).zfill(6)
                dataDict     = (userRefNo, str(uuid.uuid4()), 'individual', 'payorcorporate', getCompanyObj.email_address, str(uuid.uuid4()), str(uuid.uuid4()), 1, str(datetime.datetime.now()), str(datetime.datetime.now()))
                insertDTUser = deartimeDB.exec_SQL('insertCorpUser', dataDict, 'insert')
                if 'error' in insertDTUser:
                    messages.error(request, insertDTUser['error'])
                    deartimeDB.close()
                    return redirect(settings.HOST_ADDRESS+'/company-approval')
                else:
                    dataIndividual = (str(uuid.uuid4()), insertDTUser['lastID'], getCompanyObj.company_name.upper())
                    deartimeDB.exec_SQL('insertIndividual', dataIndividual, 'insert')

                    getCompanyObj.deartime_payerid = insertDTUser['lastID']
                    getCompanyObj.status = 'Verified'
                    getCompanyObj.save()
                    messages.success(request, 'Successfully Approved!')

                    with transaction.atomic():
                        saveMsgQueueObj = GenericLibraries.saveMessageQueue(getCompanyObj.email_address, 'CompanyApproveEmailView')
                    deartimeDB.close()
                    return redirect(settings.HOST_ADDRESS+'/company-approval')

            elif request.POST['formType'] == 'Reject':
                getCompanyObj = CorporateProfile.objects.get(id=request.POST['modalCompanyID'])
                getCompanyObj.rejected         = True
                getCompanyObj.remarks          = request.POST['rejectReason']
                getCompanyObj.updated_by       = request.user
                getCompanyObj.updated_datetime = datetime.datetime.now()

                getCompanyObj.save()

                with transaction.atomic():
                    saveMsgQueueObj = GenericLibraries.saveMessageQueue(getCompanyObj.email_address, 'CompanyRejectEmailView')
                messages.success(request, 'Application Rejected!')
                return redirect(settings.HOST_ADDRESS+'/company-approval')

            elif request.POST['formType'] == 'Defer':
                getCompanyObj = CorporateProfile.objects.get(id=request.POST['modalCompanyID'])
                getCompanyObj.remarks          = request.POST['deferReason']
                getCompanyObj.updated_by       = request.user
                getCompanyObj.updated_datetime = datetime.datetime.now()

                getCompanyObj.save()

                with transaction.atomic():
                    saveMsgQueueObj = GenericLibraries.saveMessageQueue(getCompanyObj.email_address, 'CompanyDeferEmailView')
                messages.success(request, 'Application Deferred!')
                return redirect(settings.HOST_ADDRESS+'/company-approval')

        getCompanyObj       = CorporateProfile.objects.get(id=companyID)
        getRelationships    = CompanyRelationship.objects.filter(company_id=companyID).first()
        getCompanyForms     = CorporateProfileFormAttachment.objects.filter(company_id=companyID)
        if len(list(getCompanyForms)) == 0:
            getCompanyForms = None

        if getCompanyForms:
            for i in getCompanyForms:
                if i.form_code == 'authorization-form':
                    hasAuthorizationForm = True

                if i.form_code == 'brd-form':
                    hasBrdForm = True

                if i.form_code == 'copd-form':
                    hasCopdForm = True

                if i.form_code == 'copd-form':
                    has917Form = True

                if i.form_code == 'copd-form':
                    has2478Form = True
                
                if i.form_code == 'copd-form':
                    has4958Form = True

                if i.form_code == 'copd-form':
                    haslrfcForm = True

                if i.form_code == 'copd-form':
                    has8Form = True

                if i.form_code == 'copd-form':
                    hasRosForm = True

        context = {
            'company'          : getCompanyObj,
            'relationships'    : getRelationships,
            'hasAuth'          : hasAuthorizationForm,
            'hasBrd'           : hasBrdForm,
            'hasCopd'          : hasCopdForm,
            'has917'           : has917Form,
            'has2478'          : has2478Form,
            'has4958'          : has4958Form,
            'hasLrfc'          : haslrfcForm,
            'has8'             : has8Form,
            'hasRos'           : hasRosForm,
            'forms'            : getCompanyForms,
            'path'             : request._current_scheme_host,
            'host_address'     : settings.HOST_ADDRESS
        }

        return render(request, 'CorporateApproval/ViewCompanyAccount.html', context)
    except Exception as ex:
        messages.error(request, str(ex))
        logger.error(str(ex),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/company-approval')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def EditCompanyAccountView(request, companySalt):
    try:
        decodeCompanySalt = base64.b64decode(companySalt)
        decodeASCII = decodeCompanySalt.decode('UTF-8')
        getCompanyID = decodeASCII.split('_')
        companyID = getCompanyID[1]

        getEntityType   = EntityType.objects.filter(is_active=True)
        getEnFormType   = CompanyFormType.objects.filter(is_active=True)
        entityFormData  = []
        formData        = []

        for eobj in getEntityType:
            for fobj in getEnFormType:
                if (eobj.id == fobj.entity_type_id ):
                    formData.append(fobj.form_type_id)
                    formattedForm = ','.join(formData)
                else:
                    formData = []

            formDict = {
                'entity_type_id'  : eobj.id,
                'entity_type_name': eobj.entity_name,
                'form_types'      : formattedForm,
            }
            formattedForm = ''
            entityFormData.append(formDict)

        if request.method == 'POST':
            try:    

                for key, value in request.POST.items():
                    if (key == 'entityType_post'):
                        entity_id = int(value)

                getEntityIdInfo = EntityType.objects.get(id=entity_id)

                for k, v in request.POST.items():
                    if '-form' in k and 'base64' not in k:
                        CorporateProfileFormAttachment.objects.filter(id=v).delete()

                getUserId      = CorporateProfile.objects.get(id=companyID)
                getCompanyInfo = GenericLibraries.registerCompany(request, companyID)

                for k, v in request.POST.items():
                    if k.startswith('base64'):
                        splitData = v.split('|')
                        filename = '/' + getCompanyInfo.company_name + '/' + splitData[0]
                        splitFile = splitData[0].split('.')
                        image_64_decode = base64.b64decode(splitData[1].split(',')[1])
                        form_code = k.split('base64-')[1].split('_0')[0]
                        open(settings.MEDIA_ROOT + filename, 'wb').write(image_64_decode)
                        new_attachment = CorporateProfileFormAttachment(
                            company         = getCompanyInfo,
                            form_code       = form_code,
                            attachment      = splitFile[0],
                            attachment_type = '.'+splitFile[1],
                        )
                        new_attachment.save()
                        if form_code not in ('brs-form', 'icPass-form', 'authorization-form'):
                            new_attachment.entity_type = getEntityIdInfo
                            new_attachment.save()

                CompanyRelationship.objects.filter(company_id=companyID).delete()

                saveRelationshipObj = GenericLibraries.saveCompanyRelationship(request, getCompanyInfo)
                getCompanyInfo = GenericLibraries.updateCompanySubmittedStatus(getCompanyInfo)
                
                messages.success(request, 'Changes successfully saved!')
                return redirect(settings.HOST_ADDRESS)
            except (Exception, IntegrityError) as e:
                messages.error(request, str(e))
                logger.error(str(e),extra={'username':request.user.id})
                return redirect(settings.HOST_ADDRESS+'/edit-company-account/{}'.format(companyID))

        getCompanyObj       = CorporateProfile.objects.get(id=companyID)
        getRelationships    = CompanyRelationship.objects.filter(company_id=companyID).first()
        getCompanyForms     = CorporateProfileFormAttachment.objects.filter(company_id=companyID)
        getRelationshipType = RelationshipType.objects.filter(is_active=True).exclude(id=getRelationships.relationship_type_id)

        getCompanyEntity    = CorporateProfileFormAttachment.objects.filter(company_id=companyID).first()
        if getCompanyEntity:
            getEntity           = EntityType.objects.get(id=getCompanyEntity.entity_type_id)
        else:
            getEntity       = EntityType.objects.get(id=1)
        getSQLConnection    = DearTimeDbConn()
        isConnected         = getSQLConnection.connect()
        if not isConnected:
            messages.error(request, settings.CONNECTION_LOST_MESSAGE)
            return redirect(settings.HOST_ADDRESS)

        queryCity       = dict(getSQLConnection.exec_SQL('cities', False, type='fetchall')['dset'])
        queryPost       = dict(getSQLConnection.exec_SQL('postal_codes', False, type='fetchall')['dset'])
        rawQueryState   = getSQLConnection.exec_SQL('states', False, type='fetchall')['dset']
        queryState      = []
        for data in rawQueryState:
            for i in data:
                queryState.append(i)
        getCompanySCP   = [getCompanyObj.state, getCompanyObj.city, getCompanyObj.postcode]

        calculator = PremiumCalculator()
        getCompanyObj.payment_due_date = datetime.datetime.strptime(calculator.checkPaymentDueDate(getCompanyObj.payment_due_date, getCompanyObj.id, getCompanyObj.payment_mode), "%Y-%m-%d").date()
        companyPayDueDate = getCompanyObj.payment_due_date
        DueDateDay        = companyPayDueDate.day
        DueDateMonth = ConvertNumtoMonth(int(companyPayDueDate.month))

        context = {
            'path'             : request.path_info,
            'company'          : getCompanyObj,
            'relationships'    : getRelationships,
            'entity'           : getEntity,
            'due_day'          : DueDateDay,
            'due_month'        : DueDateMonth,
            'forms'            : getCompanyForms,
            'relationship_type': getRelationshipType,
            'entity_data'      : entityFormData,
            'cities'           : queryCity,
            'postal_codes'     : queryPost,
            'states'           : queryState,
            'scp'              : getCompanySCP,
            'host_address'     : settings.HOST_ADDRESS,
            'http_host_address': settings.HTTP_HOST_ADDRESS,
            'companySalt'      : GenericLibraries().saltEncode(getCompanyObj)
        }
        getSQLConnection.close()
        return render(request, 'CorporateApproval/EditCompanyAccount.html', context)
    except Exception as ex:
        messages.error(request, 'Something went wrong!')
        logger.error(str(ex),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS)

def ConvertNumtoMonth(number):
    # To validate the number to be between 1 and 12
    if number < 1 or number > 12:
        return None

    month_mapping = {
        1: 'January',
        2: 'February',
        3: 'March',
        4: 'April',
        5: 'May',
        6: 'June',
        7: 'July',
        8: 'August',
        9: 'September',
        10: 'October',
        11: 'November',
        12: 'December',
    }

    month_name = month_mapping[number]
    return month_name
    

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def PackageListView(request, companySalt):
    try:
        decodeCompanySalt = base64.b64decode(companySalt)
        decodeASCII = decodeCompanySalt.decode('UTF-8')
        getCompanyID = decodeASCII.split('_')
        companyID = getCompanyID[1]
        getCompanyInfo          = CorporateProfile.objects.get(id=companyID)
        getCampaignPackage      = Package.objects.filter(under_campaign=getCompanyInfo.corporate_campaign_code)
        getPackageInfo          = Package.objects.filter(created_by_id=getCompanyInfo.user_id,under_campaign__isnull=True)
        getProducts             = Product.objects.all()
        getCoveragesObj         = GetCoverages()
        getMedicalPlansObj      = GetMedicalPlans()
        getFAQ                  = FAQ.objects.all()
        deartimeDB              = DearTimeDbConn()
        isConnected = deartimeDB.connect()

        if not isConnected:
            messages.error(request, settings.CONNECTION_LOST_MESSAGE)
            return redirect(settings.HOST_ADDRESS+'/member-list')
        medicalPlans = getMedicalPlansObj.getMedical(deartimeDB)
        deartimeDB.close()
        packageDict = []
        productDict = []
        productName = []
        for prd in getProducts:
            FaqQS = getFAQ.filter(product_id=prd.id)
            jsonData = {
                'product_name': prd.product_name,
                'faq': FaqQS
            }
            productDict.append(jsonData)
            productName.append(prd.product_name)
        for pkg in getPackageInfo:
            getPackageName      = pkg.package_name
            getPackageDesc      = pkg.description
            getPackageID        = pkg.id
            getProductId        = PackageProductMapping.objects.filter(package_id=pkg.id)                
            getCoverages        = getCoveragesObj.getCoverages(pkg, 2)

            if pkg.is_active == True:
                getIsActive = "Active"
            else:
                getIsActive = "Inactive"

            serializerPKG       = json.loads(serializers.serialize('json', [pkg]))
            serializerPKG[0]['fields'].update({
                'product'           : getProducts,
                'product_name'      : productName,
                'id'                : getPackageID,
                'name'              : getPackageName,
                'description'       : getPackageDesc,
                'is_active'         : getIsActive,
                'product_id'        : getProductId,
                'coverage'          : getCoverages,
            })
            packageDict.append(serializerPKG[0]['fields'])
        
        for pkg1 in getCampaignPackage:
            getPackageName      = pkg1.package_name
            getPackageDesc      = pkg1.description
            getPackageID        = pkg1.id
            getProductId        = PackageProductMapping.objects.filter(package_id=pkg1.id)                
            getCoverages        = getCoveragesObj.getCoverages(pkg1, 2)

            if pkg1.is_active == True:
                getIsActive = "Active"
            else:
                getIsActive = "Inactive"

            serializerPKG       = json.loads(serializers.serialize('json', [pkg1]))
            serializerPKG[0]['fields'].update({
                'product'           : getProducts,
                'product_name'      : productName,
                'id'                : getPackageID,
                'name'              : getPackageName,
                'description'       : getPackageDesc,
                'is_active'         : getIsActive,
                'product_id'        : getProductId,
                'coverage'          : getCoverages,
            })
            packageDict.append(serializerPKG[0]['fields'])
        packageTable = Paginator(packageDict, 10)

        context = {
            'path': request.path_info,
            'company': getCompanyInfo,
            'packageListQS': packageTable,
            'productNameQS': productDict,
            'productName'  : productName,
            'medicalPlanQS': medicalPlans,
            'product_count': len(productDict),
            'host_address' : settings.HOST_ADDRESS,
            'companySalt'  : GenericLibraries().saltEncode(getCompanyInfo),
        }
        return render(request, "CorporateApproval/PackageList.html", context)

    except Exception as ex:
        messages.error(request, str(ex))
        logger.error(str(ex),extra={'username':request.user.id})
        return render(request, "Menu/MemberList.html", context)

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def EditPackage(request, companySalt):
    try:
        decodeCompanySalt = base64.b64decode(companySalt)
        decodeASCII = decodeCompanySalt.decode('UTF-8')
        getCompanyID = decodeASCII.split('_')
        companyID = getCompanyID[1]
        user = CorporateUser.objects.get(id=request.user.id)
        productKeyFields = [prd.product_name for prd in Product.objects.filter(is_active=True)]
        coverage_dict = {}
        for k, v in request.POST.lists():
            if '_POST' in k:
                coverage_dict[k.split('_')[0]] = 0 if (v[0] == '' or v[0] == 'None') else int(v[0])
        for prd in productKeyFields:
            if not prd in coverage_dict:
                coverage_dict[prd] = 0

        if "update" in request.POST["flag"]:
            getPackages = Package.objects.filter(created_by_id=request.user.id)
            for pkg in getPackages:
                if request.POST["name"] == pkg.package_name:
                    if int(request.POST['package_id']) != pkg.id:
                        messages.error(request, "Package name already exist!")
                        return redirect(settings.HOST_ADDRESS+'/package-list/{}'.format(companySalt))

            hasCoverage = False
            for product in productKeyFields:
                if coverage_dict[product] != 0:
                    hasCoverage = True
            if not hasCoverage:
                messages.error(request, "Please select coverage amount for at least one product!")
                return redirect(settings.HOST_ADDRESS+'/package-list/{}'.format(companySalt))

            getSelectedPackage              = Package.objects.get(id=request.POST["package_id"])
            getSelectedPackage.package_name = request.POST["name"]
            getSelectedPackage.description  = None if request.POST["description"] == 'None' else request.POST["description"]
            status = False
            for key, value in request.POST.lists():
                if "status" in key:
                    status = True

            getSelectedPackage.is_active    = status
            getSelectedPackage.save()

            with transaction.atomic():
                for key,value in coverage_dict.items():
                    productObj  = Product.objects.get(product_name=key)
                    getSelectedMapping = PackageProductMapping.objects.filter(package_id=request.POST["package_id"], product=productObj)
                    if getSelectedMapping:
                        getProductPackageMapping = PackageProductMapping.objects.get(package_id=request.POST["package_id"], product=productObj)
                        getProductPackageMapping.coverage_amount = value
                        getProductPackageMapping.save()
                    else:
                        productData = PackageProductMapping(
                            coverage_amount = value,
                            package_id = request.POST["package_id"],
                            product    = productObj,
                            created_by = user
                        )
                        productData.save() 
            
            #delete coverage
            if "delete" in request.POST['flag']:
                GenericLibraries.deleteCoverage(request, request.POST["package_id"])

        #add package
        elif "add" in request.POST["flag"]:
            getPackages = Package.objects.filter(created_by_id=request.user.id)
            for pkg in getPackages:
                if request.POST["name"] == pkg.package_name:
                    messages.error(request, "Package name already exist!")
                    return redirect(settings.HOST_ADDRESS+'/package-list/{}'.format(companySalt))

            hasCoverage = False
            for product in productKeyFields:
                for k, v in request.POST.lists():
                    if coverage_dict[product] != 0:
                        hasCoverage = True
            if not hasCoverage:
                messages.error(request, "Please select coverage amount for at least one product!")
                return redirect(settings.HOST_ADDRESS+'/package-list/{}'.format(companySalt))

            # coverages rounding
            for prd in productKeyFields:
                if prd != 'Medical':
                    coverage_dict[prd] = round(coverage_dict[prd] / 1000) * 1000

            existPackage = CheckUniquePackage().check(coverage_dict, user.id)
            if existPackage:
                messages.error(request, "Package with same coverages found: {PACKAGE_NAME}".format(PACKAGE_NAME=existPackage.package_name))
                return redirect(settings.HOST_ADDRESS+'/package-list/{}'.format(companySalt))

            status = False
            for key, value in request.POST.lists():
                if "status" in key:
                    status = True

            with transaction.atomic():
                new_package  = Package (
                    package_name    = request.POST['name'],
                    description     = request.POST['description'],
                    is_active       = status,
                    created_by_id   = request.user.id
                )
                new_package.save()
                    
            with transaction.atomic():
                mappingToCreate = []
                mappingData     = []
                mappingDict     = {}

                for p in productKeyFields:
                    mappingToCreate.append(p)

                for prd in productKeyFields:
                    product  = Product.objects.get(product_name=prd)
                    mappingDict.update({
                        'is_active'       : status,
                        'coverage_amount' : coverage_dict[prd],
                        'package_id'      : new_package.id,
                        'product'         : product,
                        'created_by_id'   : request.user.id
                    })
                    mappingData.append(PackageProductMapping(**mappingDict))
                    mappingToCreate.remove(prd)

                for prd2 in mappingToCreate:
                    mappingDict.update({
                        'is_active'       : status,
                        'coverage_amount' : 0,
                        'package_id'      : new_package.id,
                        'product_id'      : Product.objects.get(product_name=prd2).id,
                        'created_by_id'   : request.user.id
                    })
                    mappingData.append(PackageProductMapping(**mappingDict))
                
                if mappingData:
                    PackageProductMapping.objects.bulk_create(mappingData)
                        
            #delete coverage
            if "delete" in request.POST['flag']:
                GenericLibraries.deleteCoverage(request, new_package)
        messages.success(request, "Successfully updated!")

    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
    return redirect(settings.HOST_ADDRESS+'/package-list/{}'.format(companySalt))

@csrf_exempt
#Not using in this stage
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def CompanyApprovalMemberListView(request, companySalt):
    try:
        decodeCompanySalt = base64.b64decode(companySalt)
        decodeASCII = decodeCompanySalt.decode('UTF-8')
        getCompanyID = decodeASCII.split('_')
        companyID = getCompanyID[1]
        getCompanyInfo    = CorporateProfile.objects.get(id=companyID)
        getMembersObj     = GetMembers()
        memberTable = getMembersObj.getMembers(corporate=getCompanyInfo)
        members = Member.objects.all().filter(corporate_id=companyID)

        invoiceMemberMapping=[]
    
        for member in members:
            orderObj = Order.objects.filter(member_id = member.id)
            if orderObj:
                order = Order.objects.filter(member_id = member.id).order_by('-created_datetime').first()
                try:
                    invoice = Invoice.objects.exclude(status='Void').get(id=order.invoice_id)
                    jsonData = {
                        'member_id': member.id,
                        'invoiceDate': invoice.created_datetime.strftime('%Y-%m-%d'),
                    }
                except Invoice.DoesNotExist:
                    # Handle the case when the invoice with the provided ID does not exist or is "Void"
                    jsonData = {
                        'member_id': member.id,
                        'invoiceDate':None,
                    }
            else:
                jsonData = {
                    'member_id':member.id,
                    'invoiceDate':None,
                }
            invoiceMemberMapping.append(jsonData)

        paymentDueMapping=[]
        for member in members:
            if member.rejected == False and member.void == False:
                jsonData = {
                    'member_id':member.id,
                    'paymentDate':getCompanyInfo.payment_due_date,
                }
            else:
                jsonData = {
                    'member_id':member.id,
                    'paymentDate':None,
                }
            paymentDueMapping.append(jsonData)

        if memberTable.object_list:
            membersCount = 0
            for mem in memberTable.object_list:
                getMember = Member.objects.get(id=mem['id'])
                if getMember.deartime_memberid:
                    getMemberADPremium = PremiumAdjustment.objects.filter(member_id=getMember.id)
                    if getMemberADPremium:
                        for  member in getMemberADPremium:
                            memberTable.object_list[membersCount]['ad_premium'] = member.ad_premium
                            memberTable.object_list[membersCount]['remarks'] = member.remarks
                
                membersCount+=1
                
        context = {
            'path'           : request.path_info,
            'company'        : getCompanyInfo,
            'premiumHolderQS': memberTable,
            'invoiceMemberMapping':invoiceMemberMapping,
            'paymentDueMapping':paymentDueMapping,
            'host_address' : settings.HOST_ADDRESS,
            'companySalt'  : GenericLibraries().saltEncode(getCompanyInfo)
        }
        
        if request.method == 'POST':
            for k, v in request.POST.lists():
                if 'searchMember' in k:
                    context['premiumHolderQS'] = getMembersObj.getMembers(corporate=getCompanyInfo, name__icontains=request.POST['searchMember'])
                    context['filterMember']    = request.POST['searchMember']
        return render(request, 'CorporateApproval/MemberList.html', context)
    except Exception as ex:
        messages.error(request, str(ex))
        logger.error(str(ex),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/company-approval')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def premiumAdjustment(request):
    try:
        if request.method == 'POST':
            member_id = request.POST['memberID']
            ADPremium = request.POST['AD_amount']
            remarks   = request.POST['remarks']
            
            ad_premium = PremiumAdjustment.objects.filter(member_id=member_id)
            if ad_premium:
                ad_premium.update(ad_premium = ADPremium)
                ad_premium.update(remarks = remarks)
            else:
                PremiumAdjustment.objects.create(member_id=member_id, ad_premium = ADPremium, remarks = remarks)
                
            messages.success(request, 'Successfully Updated!')    
            return redirect(settings.HOST_ADDRESS+'/company-approval')
        
        
    except Exception as ex:
        messages.error(request, str(ex))
        logger.error(str(ex),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/company-approval')
    
@csrf_exempt
#Not using in this stage
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def DashboardView(request):
    try:
        return redirect(settings.HOST_ADDRESS+'/member-list')
    except Exception as ex:
        messages.error(request, str(ex))
        logger.error(str(ex),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS)

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def AddMember(request):
    try:
        getUserCorporateObj = CorporateProfile.objects.get(user__id=request.user.id)
        getUserObj          = CorporateUser.objects.get(id=request.user.id) 
        packageID           = request.POST['package']
        if not packageID:
            messages.error(request, 'Please select a package.')
            return redirect(settings.HOST_ADDRESS+'/member-list')
        getPackageObj       = Package.objects.get(id=packageID)
        batch_no            = 'BN' + str(datetime.date.today().year) + str(datetime.date.today().month) + '-' + str(get_next_value(sequence_name="registration/" + str(datetime.date.today().year) + str(datetime.date.today().month))).zfill(3)
        productKeyFields    = [prd.product_name for prd in Product.objects.filter(is_active=True)]
        hasError = False
        country_code = "+60"

        # Formatting mobile number
        concat_mobile_number = request.POST['employee_countrycode'] + request.POST['employee_mobile']
        formatted_mobile_number = re.sub('[^0-9+]+', '', str(concat_mobile_number))    
        if country_code in formatted_mobile_number:
            formatted_mobile_number = formatted_mobile_number[2:]

        # Check existing member in CPFO
        checkCPFOExistMember = GenericLibraries().checkCPFOExistMember(request.POST['employee_email'], formatted_mobile_number, request.POST['employee_nric'], getUserCorporateObj)

        if not checkCPFOExistMember:
            hasError = True

        coverage_dict = {}        
        for prd in productKeyFields:
            for key, value in request.POST.lists():
                if key.split("_")[0] == packageID and "_addMemberPOST" in key:
                    if value[0] == 'None':
                        coverage_dict[prd] = 0
                    else:
                        coverage_dict[prd] = int(request.POST[packageID+"_"+prd+'_addMemberPOST'])

        # Check accident coverage not more than death coverage
        if coverage_dict['Accident'] != 0:
            if coverage_dict['Death'] == 0 or coverage_dict['Accident'] > coverage_dict['Death']:
                messages.error(request, settings.ACCIDENT_COVERAGE_MESSAGE)
                return redirect(settings.HOST_ADDRESS+'/member-list')
    
        hasCoverage = False
        for product in productKeyFields:
            if coverage_dict[product] != 0:                   
                hasCoverage = True
        if not hasCoverage:
            messages.error(request, "Please select coverage amount for at least one product!")
            return redirect(settings.HOST_ADDRESS+'/member-list')
        
        if not hasError:
            deartimeDB          = DearTimeDbConn()
            isConnected         = deartimeDB.connect()
            if not isConnected:
                messages.error(request, settings.CONNECTION_LOST_MESSAGE)
                return redirect(settings.HOST_ADDRESS+'/member-list')
            
            # Check existing member in Deartime DB
            getExistEmailOrMobile = deartimeDB.exec_SQL('validateExistingEmailMobile', {'MOBILE': formatted_mobile_number, 'EMAIL' : request.POST['employee_email'], 'NRIC': request.POST['employee_nric']}, 'fetchone')

            if getExistEmailOrMobile['dset']:

                getExistMember      = deartimeDB.exec_SQL('validateMember', {'MOBILE': formatted_mobile_number, 'EMAIL' : request.POST['employee_email'], 'NRIC': request.POST['employee_nric']}, 'fetchone')

                if not getExistMember['dset']:
                    messages.error(request, settings.EMAIL_MOBILE_EXISTING_MESSAGE)
                    return redirect(settings.HOST_ADDRESS+'/member-list')
                else:
                    memberID  = getExistMember['dset'][0]
                    # Check member under sponsored insurance
                    getSponsoredInsurance = deartimeDB.exec_SQL('getSponsoredInsurance', {'USER_ID': memberID}, 'fetchone')
                    if getSponsoredInsurance['dset']:
                        if request.POST['employee_gender'] == 'Male':
                            gender = 'him'
                            gender2 = 'his'
                        else:
                            gender = 'her'
                            gender2 = 'her'
                        messages.error(request, settings.UNDER_SPONSORED_INSURANCE.format(INSURED_NAME=getSponsoredInsurance['dset'][0], NEXT_RENEWAL_DATE=datetime.datetime.strftime(getSponsoredInsurance['dset'][1], '%d %B %Y'), GENDER=gender, GENDER2=gender2))
                        return redirect(settings.HOST_ADDRESS+'/member-list')

                    getIndividual    = deartimeDB.exec_SQL('getIndividual', {'USER_ID': memberID}, 'fetchone')

                    # Check member has active medical coverage
                    hasMedical = GenericLibraries().checkActiveMedical(deartimeDB, getIndividual, request.POST[packageID+"_Medical_addMemberPOST"])
                    if hasMedical:
                        hasError = True

            if not hasError:
                if request.POST['employment_no'] == '':
                    employment_no = '-'
                else:
                    employment_no = request.POST['employment_no']

                checkExistingPackageObj = CheckUniquePackage()
                existPackage = checkExistingPackageObj.check(coverage_dict, getUserCorporateObj.user_id)
                if existPackage:
                    getPackageObj = existPackage

                if not existPackage:
                    createPackage = GenericLibraries().addPackageData(getUserCorporateObj, getUserObj, productKeyFields, coverage_dict)

                new_member = Member(
                    batch_no            = batch_no,
                    employment_no       = employment_no,
                    name                = request.POST['employee_name'],
                    email_address       = request.POST['employee_email'],
                    mobile_no           = formatted_mobile_number,
                    nationality         = request.POST['employee_nationality'],
                    mykad               = request.POST['employee_nric'],
                    dob                 = request.POST['employee_dob'],
                    gender              = request.POST['employee_gender'],
                    submitted           = False,
                    package             = getPackageObj if existPackage else createPackage,
                    status              = 'Pending',
                    corporate_id        = getUserCorporateObj.id,
                    last_reminder       = datetime.datetime.now()
                )
                new_member.save()

                # Check prospect has been signed up as DT member
                if getExistEmailOrMobile['dset']:
                    if getExistMember['dset']:
                        new_member.deartime_memberid = memberID
                        if getExistMember['dset'][1]:
                            new_member.is_existing = True
                        new_member.save()

                    #Check prospect on sponsored insurance waiting list
                    getSponsoredInsuranceWaitingList = GenericLibraries().checkSIWaitingList(deartimeDB, memberID)

                    if getSponsoredInsuranceWaitingList:
                        new_member.si_waitinglist = True
                        new_member.siwaiting_email = True
                        new_member.save()
                        #update SI waiting list to DT database
                        updateSIConfirm = deartimeDB.exec_SQL('updateSponsoredInsuranceConfirm', {'USER_ID': memberID}, 'update')

                getNewMappings  = PackageProductMapping.objects.filter(package=new_member.package)
                if getNewMappings:
                    for nmpg in getNewMappings:   
                        new_member_product_mapping = MemberProductMapping(
                            coverage_amount     = nmpg.coverage_amount,
                            member_id           = new_member.id,
                            product             = nmpg.product,
                        )
                        new_member_product_mapping.save()
                getMemberObj = Member.objects.get(id=new_member.id)
                calculator = PremiumCalculator()
                quotationDict = calculator.calculate_quotation(deartimeDB, getMemberObj, getUserCorporateObj.payment_mode)
                getMemberObj.quotation_premium = quotationDict
                getMemberObj.tentative_premium = quotationDict
                getMemberObj.save()
                messages.success(request, 'Successfully added!')
            else:
                messages.error(request, settings.MEDICAL_EXISTING_MESSAGE)
            deartimeDB.close()
        else:
            messages.error(request, settings.EMAIL_MOBILE_EXISTING_MESSAGE)
    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
    return redirect(settings.HOST_ADDRESS+'/member-list')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def EditMember(request):
    try:
        getUserCorporateObj = CorporateProfile.objects.get(user__id=request.user.id)
        getMemberObj        = Member.objects.get(id=request.POST['member_id'])
        packageID           = request.POST['package']
        getPackageObj       = Package.objects.get(id=packageID)
        getPackageProductMapping = PackageProductMapping.objects.filter(created_by_id=request.user.id, package_id=getMemberObj.package_id)
        productKeyFields    = [prd.product_name for prd in Product.objects.filter(is_active=True)]
        hasError = False

        if int(request.POST[packageID+"_Accident_editMemberPOST"]) != 0:
            if int(request.POST[packageID+"_Death_editMemberPOST"]) == 0 or int(request.POST[packageID+"_Accident_editMemberPOST"]) > int(request.POST[packageID+"_Death_editMemberPOST"]):
                messages.error(request, settings.ACCIDENT_COVERAGE_MESSAGE)
                return redirect(settings.HOST_ADDRESS+'/member-list')

        hasCoverage = False
        for product in productKeyFields:
            for k, v in request.POST.lists():
                if (packageID+"_"+product+'_editMemberPOST' in k) and int(request.POST[packageID+"_"+product+'_editMemberPOST']) != 0:
                    hasCoverage = True
        if not hasCoverage:
            messages.error(request, "Please select coverage amount for at least one product!")
            return redirect(settings.HOST_ADDRESS+'/member-list')

        # Save to Deartime DB
        deartimeDB          = DearTimeDbConn()
        isConnected         = deartimeDB.connect()
        if not isConnected:
            messages.error(request, settings.CONNECTION_LOST_MESSAGE)
            return redirect(settings.HOST_ADDRESS+'/member-list')

        # Check existing nric
        if request.POST['employee_nric'] != getMemberObj.mykad:
            exist_nric = deartimeDB.exec_SQL('getIndividualNRIC', {'NRIC': request.POST['employee_nric']}, 'fetchone')
            check_nric = Member.objects.filter(mykad=request.POST['employee_nric'])
            if exist_nric['dset'] or check_nric:
                messages.error(request, 'NRIC has existed.')
                return redirect(settings.HOST_ADDRESS+'/member-list')

        # Check existing member in Deartime DB
        getExistEmailOrMobile = deartimeDB.exec_SQL('validateExistingEmailMobile', {'MOBILE': getMemberObj.mobile_no, 'EMAIL' : getMemberObj.email_address, 'NRIC': getMemberObj.mykad}, 'fetchone')

        if getExistEmailOrMobile['dset']:

            getExistMember    = deartimeDB.exec_SQL('validateMember', {'MOBILE': getMemberObj.mobile_no, 'EMAIL' : getMemberObj.email_address, 'NRIC': getMemberObj.mykad}, 'fetchone')

            if not getExistMember['dset']:
                messages.error(request, settings.EMAIL_MOBILE_EXISTING_MESSAGE)
                return redirect(settings.HOST_ADDRESS+'/member-list')
            else:
                updateUserDict = {'EMAIL': getMemberObj.email_address, 'DEARTIME_MEMBERID': getMemberObj.deartime_memberid}
                updateDTUser = deartimeDB.exec_SQL('updateMemberEmail', updateUserDict, 'update')
                if 'error' in updateDTUser:
                    hasError = True    
                else:
                    updateIndividual = {'NAME': request.POST['employee_name'].upper(), 'NRIC': request.POST['employee_nric'], 'MOBILE': getMemberObj.mobile_no, 'GENDER': request.POST['employee_gender'], 'DOB': request.POST['employee_dob'], 'NATIONALITY': request.POST['employee_nationality'], 'DEARTIME_MEMBERID': getMemberObj.deartime_memberid}
                    deartimeDB.exec_SQL('updateIndividualMember', updateIndividual, 'update')

                    getMedicalPlans = GetMedicalPlans()
                    isConnected = getMedicalPlans.connect()
                    if not isConnected:
                        messages.error(request, settings.CONNECTION_LOST_MESSAGE)
                        return redirect(settings.HOST_ADDRESS+'/member-list')
                    medical_deductibles = getMedicalPlans.getMedical(deartimeDB)
        if not hasError:
            new_coverages ={}
            for prds in productKeyFields:
                for key, value in request.POST.lists():
                    if 'editMemberPOST' in key:
                        if key.split("_")[0] == packageID and key.split("_")[1].upper() == prds.upper():
                            new_coverages[prds] = int(value[0])

            for prodMap in getPackageProductMapping:
                getExistingMemberPrdMap = MemberProductMapping.objects.get(member_id=request.POST['member_id'], product_id=prodMap.product_id, is_terminated=False, is_renewal=False)
                getExistingMemberPrdMap.is_terminated = True
                getExistingMemberPrdMap.updated_datetime = datetime.datetime.now()
                getExistingMemberPrdMap.save()
                if getExistingMemberPrdMap.deartime_coverageid:
                    getProductID = deartimeDB.exec_SQL('getProductObj', {'PRD_NAME': prodMap.product.product_name.title()}, 'fetchone')
                    individualID = deartimeDB.exec_SQL('getIndividual', {'USER_ID': getMemberObj.deartime_memberid}, 'fetchone')
                    deartimeDB.exec_SQL('updateSingleTerminateCoverage', {'OWNER_ID': individualID['dset'][0], 'PAYER_ID': getUserCorporateObj.deartime_payerid, 'PRODUCT_ID': getProductID['dset'][0]}, 'update')

            for prd, cvg in new_coverages.items():
                getProductObj  = Product.objects.get(product_name=prd)
                new_member_product_mapping = MemberProductMapping(
                    coverage_amount     = cvg,
                    member_id           = request.POST['member_id'],
                    product             = getProductObj,
                )
                new_member_product_mapping.save()

            checkExistingPackageObj = CheckUniquePackage()
            productToCheck = {}
            for prd in productKeyFields:
                productToCheck[prd] = int(request.POST[packageID+"_"+prd+"_editMemberPOST"])                
            existPackage = checkExistingPackageObj.check(productToCheck, getUserCorporateObj.user_id)

            if existPackage:
                getPackageObj = existPackage

            if not existPackage:
                createPackage = GenericLibraries.editPackageData(getUserCorporateObj, request, productKeyFields, packageID)

            calculator = PremiumCalculator()
            quotationDict = calculator.calculate_quotation(deartimeDB, getMemberObj, getUserCorporateObj.payment_mode)
            getMemberObj.quotation_premium = quotationDict
            getMemberObj.tentative_premium = quotationDict

            if request.POST['employment_no'] == '':
                employment_no = '-'
            else:
                employment_no = request.POST['employment_no']

            getMemberObj.employment_no   = employment_no
            getMemberObj.name            = request.POST['employee_name']
            getMemberObj.email_address   = getMemberObj.email_address
            getMemberObj.mobile_no       = getMemberObj.mobile_no
            getMemberObj.nationality     = request.POST['employee_nationality']
            getMemberObj.mykad           = request.POST['employee_nric']
            getMemberObj.dob             = request.POST['employee_dob']
            getMemberObj.gender          = request.POST['employee_gender']
            getMemberObj.package         = getPackageObj if existPackage else createPackage
            getMemberObj.status          = 'Pending'
            getMemberObj.save()

            messages.success(request, 'Successfully updated!')
        else:
            messages.error(request, updateDTUser['error'])
        deartimeDB.close()
    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
    return redirect(settings.HOST_ADDRESS+'/member-list')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def MemberListView(request):
    try:
        if request.method == 'POST':
            members = {k: v for k, v in request.POST.items() if k.startswith('member')}
            if members:
                for memberKey, memberValue in members.items():
                    # Save to Deartime DB
                    deartimeDB          = DearTimeDbConn()
                    isConnected         = deartimeDB.connect()
                    if not isConnected:
                        messages.error(request, settings.CONNECTION_LOST_MESSAGE)
                        return redirect(settings.HOST_ADDRESS+'/member-list')
                    getMemberObj = Member.objects.get(id=memberValue)
                    
                    checkExistingMember = deartimeDB.exec_SQL('validateMember', {'MOBILE': getMemberObj.mobile_no, 'EMAIL' : getMemberObj.email_address, 'NRIC' : getMemberObj.mykad}, 'fetchone')
                    if checkExistingMember['dset']:
                        individualID  = deartimeDB.exec_SQL('getIndividual', {'USER_ID': checkExistingMember['dset'][0]}, 'fetchone')
                        getMemberObj.deartime_memberid = checkExistingMember['dset'][0]
                        if checkExistingMember['dset'][1]:
                            getMemberObj.is_existing = True       
                    else:
                        insertUser   = GenericLibraries().insertUserToDTDB(deartimeDB, memberValue)
                        individualID = deartimeDB.exec_SQL('getIndividual', {'USER_ID': insertUser}, 'fetchone')
                        getMemberObj.deartime_memberid = insertUser
                    getMemberProductMappingObj =  MemberProductMapping.objects.filter(member_id=memberValue, is_terminated=False, is_renewal=False).exclude(deartime_coverageid__isnull=True)
                    if getMemberProductMappingObj:
                        for memberProduct in getMemberProductMappingObj:
                            getExistCoverage = deartimeDB.exec_SQL('validateCoverage', {'USER_ID': getMemberObj.deartime_memberid, 'PRODUCT_NAME':memberProduct.product.product_name.title()}, 'fetchone')
                            if getExistCoverage['dset']:
                                updateCoverageDate = deartimeDB.exec_SQL('updateCoverageDateIncreaseUnpaid', {'UPDATED_DATE': str(datetime.datetime.now()), 'COVERAGE_ID':memberProduct.deartime_coverageid}, 'update')
                            else:
                                updateCoverageDate = deartimeDB.exec_SQL('updateCoverageDate', {'UPDATED_DATE': str(datetime.datetime.now()), 'COVERAGE_ID':memberProduct.deartime_coverageid}, 'update')
                            memberProduct.updated_datetime = datetime.datetime.now()
                            memberProduct.save()
                    else:
                        insertCoverage = GenericLibraries().insertCoverageToDTDB(deartimeDB, memberValue, individualID)
                    getMemberObj.status = 'Pending Acceptance'
                    getMemberObj.sendinvitation_datetime = datetime.datetime.now()
                    getMemberObj.save()
                    
                    saveMessageQueue  = MessagingQueue(
                        email_address = getMemberObj.email_address,
                        module        = 'MemberInvitationView'
                    )
                    saveMessageQueue.save()
                    
                    if getMemberObj.siwaiting_email: 
                        saveAppMessageQueue = MessagingQueue(
                            email_address = getMemberObj.email_address,
                            module        = 'SIMemberInvitationApp'
                        )
                    else:
                        saveAppMessageQueue = MessagingQueue(
                            email_address = getMemberObj.email_address,
                            module        = 'MemberInvitationApp'
                        )
                    saveAppMessageQueue.save()
                    if not getMemberObj.is_existing and getMemberObj.deartime_memberid:
                        individualID      = deartimeDB.exec_SQL('getIndividual', {'USER_ID': getMemberObj.deartime_memberid}, 'fetchone')
                        getVerifiedEKYC   = deartimeDB.exec_SQL('getVerifiedEKYC', {'INDIVIDUAL_ID': individualID['dset'][0]}, 'fetchone')
                        if getVerifiedEKYC['dset']:
                            if getVerifiedEKYC['dset'][0] != 'Accepted':
                                saveEKYCMessageQueue = MessagingQueue(
                                    email_address = getMemberObj.email_address,
                                    module        = 'MemberEkycApp'
                                )
                                saveEKYCMessageQueue.save()
                        else:
                            saveEKYCMessageQueue = MessagingQueue(
                                email_address = getMemberObj.email_address,
                                module        = 'MemberEkycApp'
                            )
                            saveEKYCMessageQueue.save()
                    deartimeDB.close()
                messages.success(request, 'Successfully Sent!')
                return redirect(settings.HOST_ADDRESS+'/member-list')
        
        deartimeDB                     = DearTimeDbConn()
        isConnected                    = deartimeDB.connect()
        if not isConnected:
            messages.error(request, settings.CONNECTION_LOST_MESSAGE)
            return redirect(settings.HOST_ADDRESS)
        calculator = PremiumCalculator()
        genericLibraries = GenericLibraries()
        getCompanyObj                  = CorporateProfile.objects.get(user_id=request.user.id)
        #genericLibraries.memberAcceptedStatus(deartimeDB, getCompanyObj)
        genericLibraries.activeSponsoredInsurance(deartimeDB, getCompanyObj)
        genericLibraries.coverageStatus(deartimeDB, getCompanyObj)
        genericLibraries.checkPassMedicalSurvey(deartimeDB, getCompanyObj)
        genericLibraries.checkChangePaymentMode(getCompanyObj)
        getCompanyObj.payment_due_date = datetime.datetime.strptime(calculator.checkPaymentDueDate(getCompanyObj.payment_due_date, getCompanyObj.id, getCompanyObj.payment_mode), "%Y-%m-%d").date()
        getPackageQS                   = Package.objects.filter(created_by_id=request.user.id,under_campaign__isnull = True)
        getCampaignPackages            = Package.objects.filter(under_campaign=getCompanyObj.corporate_campaign_code)
        combinedPackages               = list(getPackageQS) + list(getCampaignPackages)
        # get campaign records and convert to list
        getCampaignRecords             = deartimeDB.exec_SQL('getCampaignList',{},'fetchall')
        campaignList                   = (record[0] for record in getCampaignRecords['dset'])
        getProductQS                   = Product.objects.filter(is_active=True)
        getCoveragesObj                = GetCoverages()
        getMembersObj                  = GetMembers()
        getMedicalObj                  = GetMedicalPlans()
        
        # new trigger to fetch additional records from vapor
        getAdditions = deartimeDB.exec_SQL('getVoucherMembers',{'PAYER_ID': getCompanyObj.deartime_payerid},'fetchall')
        productKeyFields = [prd.product_name for prd in Product.objects.filter(is_active=True)]
        batch_no   = 'BN' + str(datetime.date.today().year) + str(datetime.date.today().month) + '-' + str(get_next_value(sequence_name="registration/" + str(datetime.date.today().year) + str(datetime.date.today().month))).zfill(3)
        
        length = len(getAdditions['dset'])
        # starts a for loop to get all the records store in a new list
        for i in range(length):
            # get latest member id and plus one as new id 
            currentMemberId = Member.objects.aggregate(Max('id'))['id__max']
            newMemberId = str(int(currentMemberId) + 1)
            # not done
            for member in getAdditions['dset']:  
                coverageFields = [
                    {'Death':0},
                    {'Disability':0},
                    {'Critical':0},
                    {'Accident':0},
                    {'Medical':0}
                ]  
                # check exist member
                checkCPFOExistMember = GenericLibraries().checkCPFOExistMember(member[2], member[4], member[0], member[3], getCompanyObj)            
                
                # if dont have member, only run this
                if checkCPFOExistMember:
                    # get all the coverages based on the campaign and member nric
                    all_coverages = deartimeDB.exec_SQL('getAdditionCoverage',{'USER_NRIC':member[0], 'CAMPAIGN':member[3]},'fetchall')
                    employment_no = ''
                    try:
                        getMember = Member.objects.filter(mykad=member[0]).first()
                        employment_no = getMember.employment_no
                    except Member.DoesNotExist:
                        getMember = None
                        employment_no = '-'

                    coverage_dict = {coverage[0]: coverage[1] for coverage in all_coverages['dset']}
                    print(coverage_dict)

                    for coverage in coverageFields:
                        for coverage_type in coverage:
                            if coverage_type in coverage_dict:
                                # Set the amount
                                coverage[coverage_type] = coverage_dict[coverage_type] 
        
                    for prd in productKeyFields:
                        if prd != 'Medical':
                            if prd in coverageFields:
                                coverage_dict[prd] = int(GenericLibraries().round_half_up(coverage_dict[prd] / 1000) * 1000)

                    checkExistPackageObj = CheckUniquePackage()
                    productToCheck = {}
                    for prd in productKeyFields:
                        productToCheck[prd] = int(coverage_dict.get(prd, 0)) 

                    existPackage = checkExistPackageObj.check(productToCheck, getCompanyObj.user_id, getCompanyObj.corporate_campaign_code)
                    print(existPackage)
                    new_package = ''
                    
                    if not existPackage:
                        createPackage = Package(
                            package_name='Customized-' + str(get_next_value(sequence_name=getCompanyObj.company_name + "_package/")).zfill(3),
                            created_by=request.user
                        )
                        createPackage.save()

                        productData = []
                        for prd in productKeyFields:
                            getProductObj = Product.objects.filter(product_name__icontains=prd).first()
                            if getProductObj:
                                coverage_amount = coverage_dict.get(prd, 0)  # Use the updated coverage fields
                                productData.append(
                                    PackageProductMapping(
                                        package=createPackage,
                                        product=getProductObj,
                                        coverage_amount=coverage_amount,
                                        created_by=request.user
                                    )
                                )
                                new_package = createPackage
                    else:
                        new_package = existPackage.id
                        # new member list        
                        new_member_data = {
                            'id': newMemberId,
                            'batch_no': batch_no,
                            'employment_no':employment_no,
                            'name': member[1],
                            'email_address':member[2],
                            'mobile_no':member[4],
                            'nationality':member[5],
                            'mykad': member[0],
                            'passport':None,
                            'dob':member[6],
                            'gender':member[7],
                            'tentative_premium':0.00,
                            'submitted':0,
                            'medical_survey':1,
                            'paid':0,
                            'status':'Accept',
                            'void':0,
                            'deartime_memberid':member[8],
                            'last_reminder':None,
                            'rejected':0,
                            'created_datetime':datetime.datetime.now(),
                            'is_existing':'',
                            'reminder_count': None,
                            'generated_invoice':0,
                            'corporate_id': getCompanyObj.id,
                            'package_id':new_package,
                            'is_deleted':0,
                            'quotation_premium': 0.00,
                            'rejected_reason':None,
                            'si_waitinglist':0,
                            'siwaiting_email':0,
                            'true_premium':0.00,
                            'invoice_reminder_count':0,
                            'sendinvitation_datetime':None,
                            'read_datetime':None,
                            'updated_datetime':datetime.datetime.now(),
                            'renew':0,
                            'campaign_code': member[3]
                        }   
                        mappedData = Member(**new_member_data)
                        
                        if mappedData:
                            with transaction.atomic():
                                mappedData.save()
                                getMember = mappedData

                        calculator = PremiumCalculator()
                        memberProductData   = []
                        if getMember:
                            getNewMappings      = PackageProductMapping.objects.filter(package=getMember.package)
                            if getNewMappings:
                                for nmpg in getNewMappings:
                                    memberProductDict = {
                                        'member_id'             : getMember.id,
                                        'product_id'            : nmpg.product_id,
                                        'coverage_amount'       : nmpg.coverage_amount
                                    }
                                    memberProductData.append(
                                        MemberProductMapping(**memberProductDict)
                                    )
                                    if memberProductData:
                                        with transaction.atomic():
                                            getNewMemberProductMappings = MemberProductMapping.objects.bulk_create(memberProductData)
                                            memberProductData = []
                                
                            quotationDict = calculator.calculate_quotation(deartimeDB, getMember, getCompanyObj.payment_mode, campaignCode=True)
                            getMember.quotation_premium = quotationDict
                            getMember.tentative_premium = quotationDict
                            getMember.save()
                        
        
        membersTable    = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=False, rejected=False, void=False, generated_invoice=False, is_deleted=False)
        pendingTable    = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=True, rejected=False, void=False, generated_invoice=False, is_deleted=False, renew = False)
        renewalTable    = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=True, rejected=False, void=False, generated_invoice=False, is_deleted=False, renew = True)
        paidTable       = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=True, rejected=False, void=False, is_deleted=False)
        rejectedTable   = getMembersObj.getMembers(corporate_id=getCompanyObj.id, rejected=True, void=False, is_deleted=False)
        terminatedTable = getMembersObj.getMembers(corporate_id=getCompanyObj.id, void=True, is_deleted=False)
        
        # filter function to filter all the tables, to be integrate and enhance
        campaign_code_filter = request.POST.get('searchCampaignCode', '')
        if request.method == 'POST':
            if campaign_code_filter == 'null': 
                membersTable    = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=False, rejected=False, void=False, generated_invoice=False, is_deleted=False, campaign_code__isnull=True)
                pendingTable    = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=True, rejected=False, void=False, generated_invoice=False, is_deleted=False, renew = False, campaign_code__isnull=True)
                renewalTable    = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=True, rejected=False, void=False, generated_invoice=False, is_deleted=False, renew = True, campaign_code__isnull=True)
                paidTable       = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=True, rejected=False, void=False, is_deleted=False, campaign_code__isnull=True)
                rejectedTable   = getMembersObj.getMembers(corporate_id=getCompanyObj.id, rejected=True, void=False, is_deleted=False, campaign_code__isnull=True)
                terminatedTable = getMembersObj.getMembers(corporate_id=getCompanyObj.id, void=True, is_deleted=False, campaign_code__isnull=True)
                
            elif campaign_code_filter == 'not_null': 
                membersTable    = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=False, rejected=False, void=False, generated_invoice=False, is_deleted=False, campaign_code__isnull=False)
                pendingTable    = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=True, rejected=False, void=False, generated_invoice=False, is_deleted=False, renew = False, campaign_code__isnull=False)
                renewalTable    = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=True, rejected=False, void=False, generated_invoice=False, is_deleted=False, renew = True, campaign_code__isnull=False)
                paidTable       = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=True, rejected=False, void=False, is_deleted=False, campaign_code__isnull=False)
                rejectedTable   = getMembersObj.getMembers(corporate_id=getCompanyObj.id, rejected=True, void=False, is_deleted=False, campaign_code__isnull=False)
                terminatedTable = getMembersObj.getMembers(corporate_id=getCompanyObj.id, void=True, is_deleted=False, campaign_code__isnull=False)
                
            elif campaign_code_filter:
                membersTable    = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=False, rejected=False, void=False, generated_invoice=False, is_deleted=False, campaign_code=campaign_code_filter)
                pendingTable    = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=True, rejected=False, void=False, generated_invoice=False, is_deleted=False, renew = False, campaign_code=campaign_code_filter)
                renewalTable    = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=True, rejected=False, void=False, generated_invoice=False, is_deleted=False, renew = True, campaign_code=campaign_code_filter)
                paidTable       = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=True, rejected=False, void=False, is_deleted=False, campaign_code=campaign_code_filter)
                rejectedTable   = getMembersObj.getMembers(corporate_id=getCompanyObj.id, rejected=True, void=False, is_deleted=False, campaign_code=campaign_code_filter)
                terminatedTable = getMembersObj.getMembers(corporate_id=getCompanyObj.id, void=True, is_deleted=False, campaign_code=campaign_code_filter)
                    
            else: 
                membersTable    = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=False, rejected=False, void=False, generated_invoice=False, is_deleted=False)
                pendingTable    = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=True, rejected=False, void=False, generated_invoice=False, is_deleted=False, renew = False)
                renewalTable    = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=True, rejected=False, void=False, generated_invoice=False, is_deleted=False, renew = True)
                paidTable       = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=True, rejected=False, void=False, is_deleted=False)
                rejectedTable   = getMembersObj.getMembers(corporate_id=getCompanyObj.id, rejected=True, void=False, is_deleted=False)
                terminatedTable = getMembersObj.getMembers(corporate_id=getCompanyObj.id, void=True, is_deleted=False)

        currentDate = GenericLibraries().currentDateTesting(getCompanyObj.id)
        cc = ''
        exist = False
        
        if membersTable.object_list:
            membersCount = 0
            for mem in membersTable.object_list:
                getMember = Member.objects.get(id=mem['id'])
                if getMember.deartime_memberid:
                    getMemberDetails = deartimeDB.exec_SQL('getProspectDetails', {'USER_ID': getMember.deartime_memberid}, 'fetchone')
                    if getMemberDetails['dset']:
                        getMember.dob = getMemberDetails['dset'][2]
                        getMember.gender = getMemberDetails['dset'][3]
                        getMember.save()
                membersTable.object_list[membersCount]['dob'] = getMember.dob.strftime("%Y-%m-%d")
                membersTable.object_list[membersCount]['gender'] = getMember.gender
                if (getMember.campaign_code):
                    cc = True
                membersTable.object_list[membersCount]['quotation_premium'] = calculator.calculate_quotation(deartimeDB, getMember, getCompanyObj.payment_mode, has_CampaignCode=cc)
                if not getMember.quotation_premium:
                    getMember.quotation_premium = calculator.calculate_quotation(deartimeDB, getMember, getCompanyObj.payment_mode, has_CampaignCode=cc)
                    getMember.save()                
                if mem['employment_no'] == '-':
                    mem['employment_no'] = ''
                membersCount+=1
        
        pendingCount = 0
        if pendingTable.object_list:
            for pendingMem in pendingTable.object_list:
                getMember = Member.objects.get(id=pendingMem['id'])
                #getMemberProductMapping = MemberProductMapping.objects.filter(member_id=pendingMem['id'], is_terminated=False).exclude(deartime_coverageid__isnull=True)
                getMemberProductMapping = MemberProductMapping.objects.filter(member_id=pendingMem['id'], is_terminated=False, is_renewal=False)
                getMemberDetails = deartimeDB.exec_SQL('getProspectDetails', {'USER_ID': getMember.deartime_memberid}, 'fetchone')
                getMember.dob = getMemberDetails['dset'][2]
                getMember.gender = getMemberDetails['dset'][3]
                getMember.save()
                getMemberADPremium = PremiumAdjustment.objects.filter(member_id=getMember.id)
                if getMemberADPremium:
                    for  member in getMemberADPremium:
                        pendingTable.object_list[pendingCount]['ad_premium'] = member.ad_premium
                        pendingTable.object_list[pendingCount]['remark'] = member.remarks
                pendingTable.object_list[pendingCount]['dob'] = getMember.dob.strftime("%Y-%m-%d")
                pendingTable.object_list[pendingCount]['gender'] = getMember.gender
                if (getMember.campaign_code):
                    cc = True
                    
                if (getMember.is_existing):
                    exist = True
                    
                pendingTable.object_list[pendingCount]['tentative_premium'] = calculator.calculate_premium(pendingMem['deartime_memberid'], getCompanyObj.id, 'total', getMemberProductMapping, deartimeDB, memberCampaignCode=cc, memberExisting=exist)
                getMember.tentative_premium = calculator.calculate_premium(pendingMem['deartime_memberid'], getCompanyObj.id, 'total', getMemberProductMapping, deartimeDB, memberCampaignCode=cc,memberExisting=exist)
                getMember.save()
                pendingCount+=1
        
        if paidTable.object_list:
            paidCount = 0
            for paidMem in paidTable.object_list:
                paidTable.object_list[paidCount]['preferredDate'] = currentDate.strftime("%Y-%m-%d")
                paidCount+=1

        # bring formatted value to select all function 
        currentDate = currentDate.strftime("%Y-%m-%d")         

        if renewalTable.object_list:
            renewalCount = 0
            for renewalMem in renewalTable.object_list:
                getMember = Member.objects.get(id=renewalMem['id'])
                getMemberProductMapping = MemberProductMapping.objects.filter(member_id=renewalMem['id'], is_terminated=False, is_renewal=True)
                getMemberDetails = deartimeDB.exec_SQL('getProspectDetails', {'USER_ID': getMember.deartime_memberid}, 'fetchone')
                getMember.dob = getMemberDetails['dset'][2]
                getMember.gender = getMemberDetails['dset'][3]
                getMember.save()
                renewalTable.object_list[renewalCount]['dob'] = getMember.dob.strftime("%Y-%m-%d")
                renewalTable.object_list[renewalCount]['gender'] = getMember.gender
                if (getMember.campaign_code):
                    cc = True
                renewalTable.object_list[renewalCount]['tentative_premium'] = calculator.calculate_premium(renewalMem['deartime_memberid'], getCompanyObj.id, 'total', getMemberProductMapping, deartimeDB, renewal=True, old_lrd = True, memberCampaignCode=cc, memberExisting=exist)
                getMember.tentative_premium = calculator.calculate_premium(renewalMem['deartime_memberid'], getCompanyObj.id, 'total', getMemberProductMapping, deartimeDB, renewal=True, old_lrd = True, memberCampaignCode=cc, memberExisting=exist)
                getMember.save()
                renewalCount+=1
         
                
        if paidTable.object_list:
            for paidMem in paidTable.object_list:
                getMember = Member.objects.get(id=paidMem['id'])
                if (PremiumAdjustment.objects.filter(member_id=paidMem['id'], paid = False, generated_invoice = False, void = False).exists()):
                    pendingTable.object_list.append(paidMem)
                    getMemberADPremium = PremiumAdjustment.objects.get(member_id=paidMem['id'], paid = False, generated_invoice = False, void = False)
                    pendingTable.object_list[pendingCount]['ad_premium'] = getMemberADPremium.ad_premium
                    pendingTable.object_list[pendingCount]['remark'] = getMemberADPremium.remarks
                    pendingTable.object_list[pendingCount]['dob'] = getMember.dob.strftime("%Y-%m-%d")
                    pendingTable.object_list[pendingCount]['gender'] = getMember.gender
                    pendingTable.object_list[pendingCount]['tentative_premium'] = getMember.tentative_premium
                    pendingCount+=1
                    
        packageProductMappingList = []
        
        for package in combinedPackages:
            getCoverages4 = getCoveragesObj.getCoverages(package, 2)
            package_data = {
                'package_id'   : package.id,
                'package_name' : package.package_name,
                'product_list' : getCoverages4
            }
            packageProductMappingList.append(package_data)
        
        activePackage = False
        for package in combinedPackages:
             # check if there is any active packages
            if package.is_active:
                activePackage = True
                break    

        medical_coverage = getMedicalObj.getMedical(deartimeDB)
        
        getComIndividual  = deartimeDB.exec_SQL('getIndividual', {'USER_ID': getCompanyObj.deartime_payerid}, 'fetchone')
        getBankAccount  = deartimeDB.exec_SQL('validateBankAccount', {'OWNER_ID': getComIndividual['dset'][0]}, 'fetchone')

        first_payment_on = None 
        gracePeriod = 0
        activeMemberRenewal = Member.objects.filter(status='Active',corporate_id=getCompanyObj.id, void=False, is_deleted=False)
        today = datetime.datetime.today()
        for member in activeMemberRenewal:
            getMemberIndividualID = deartimeDB.exec_SQL('getIndividual', {'USER_ID': member.deartime_memberid}, 'fetchone')
            getPayerID = CorporateProfile.objects.get(id=member.corporate_id)
            memberCoverage = deartimeDB.exec_SQL('getCoveragesDates', {'OWNER_ID': getMemberIndividualID['dset'][0], 'PAYER_ID': getPayerID.deartime_payerid}, 'fetchone')
            firstIndex = memberCoverage['dcolname'].index('first_payment_on')
            first_payment_on = memberCoverage['dset'][firstIndex]
            
            timeDifference = today - first_payment_on
            daysDifference = timeDifference.days
            yearsDifference = daysDifference / 365
            yearsDifference = int(yearsDifference)
            # Set the grace period based on the years difference
            if yearsDifference >= 2:
                gracePeriod = -90
            else:
                gracePeriod = -30

        bankExist = False
        if getBankAccount['dset']:
            bankExist = True
           
        context = {
            'path'            : request.path_info,
            'company'         : getCompanyObj,
            'premiumHolderQS' : membersTable,
            'payment'         : pendingTable,
            'renewal'         : renewalTable,
            'paid'            : paidTable,
            'rejected'        : rejectedTable,
            'terminated'      : terminatedTable,
            'packageQS'       : getPackageQS,
            'activePackage'   : activePackage,
            'product'         : getProductQS,
            'packageProductQS': packageProductMappingList,
            'medical_coverage': medical_coverage,
            'bankExist'       : bankExist,
            'gracePeriod'     : gracePeriod,
            'currentDate'     : currentDate,
            'host_address'    : settings.HOST_ADDRESS,
            'companySalt'     : GenericLibraries().saltEncode(getCompanyObj),
            'campaign_code_filter' : campaign_code_filter,
            'underCampaign'   : getCampaignPackages,
            'campaignList'    : campaignList,
        }
                    
        if request.method == 'POST':
            for K2, V2 in request.POST.lists():
                if "formtype" in K2 and V2 != 'invitation':
                    filterTable = K2.removeprefix("formtype")
                    context['filter'+filterTable] = request.POST['search'+filterTable]
                    #get the value from the select option field using POST
                    campaign_code_filter = request.POST.get('searchCampaignCode', '')
                    if filterTable == 'Member':
                        context['premiumHolderQS']  = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=False, rejected=False, void=False, generated_invoice=False, is_deleted=False, name__icontains=request.POST['search'+filterTable])
                    elif filterTable == 'Pending':               
                        context['payment']          = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=False, medical_survey=True, rejected=False, void=False, generated_invoice=False, is_deleted=False, name__icontains=request.POST['search'+filterTable])
                    elif filterTable == 'Paid':
                        context['paid']             = getMembersObj.getMembers(corporate_id=getCompanyObj.id, paid=True, rejected=False, void=False, is_deleted=False, name__icontains=request.POST['search'+filterTable])
                    elif filterTable == 'Rejected':
                        context['rejected']         = getMembersObj.getMembers(corporate_id=getCompanyObj.id, rejected=True, is_deleted=False, name__icontains=request.POST['search'+filterTable])
                    elif filterTable == 'Terminated':
                        context['terminated']       = getMembersObj.getMembers(corporate_id=getCompanyObj.id, void=True, is_deleted=False, name__icontains=request.POST['search'+filterTable])
                
                    # if filterTable == 'Pending' and 'payment' in context:
                    #     if campaign_code_filter == 'not_null':
                    #         context['payment'] = context['payment'].exclude(campaign_code__isnull=True)
                    #     elif campaign_code_filter == 'null':
                    #         context['payment'] = context['payment'].filter(campaign_code__isnull=True)
                        
        deartimeDB.close()
        return render(request, 'Menu/MemberList.html', context)
    except Exception as ex:
        messages.error(request, str(ex))
        logger.error(str(ex),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS)

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def memberRenewal(request):
    try:
        if request.method == 'POST':
             for key, value in request.POST.lists():
                if "paidmember" in key:
                    memberID = int(key.removeprefix('paidmember').removesuffix("_POST"))
                    getMember = Member.objects.get(id=memberID)
                    getMember.paid = False
                    getMember.generated_invoice = False
                    getMember.renew = True
                    getMember.status = 'Pending Payment'
                    getMember.save()
    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/member-list')
    return redirect(settings.HOST_ADDRESS+'/member-list')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def FailedUploadListView(request, companySalt):
    try:
        decodeCompanySalt = base64.b64decode(companySalt)
        decodeASCII = decodeCompanySalt.decode('UTF-8')
        getCompanyID = decodeASCII.split('_')
        companyID = getCompanyID[1]
        getUserCorporateObj = CorporateProfile.objects.get(id=companyID)
        context = {
            'company': getUserCorporateObj,
            'host_address' : settings.HOST_ADDRESS,
            'companySalt'  : GenericLibraries().saltEncode(getUserCorporateObj)
        }
        fileDir = os.path.join(settings.BASE_DIR, "media\\" + getUserCorporateObj.company_name + "\\FailUploadMember").replace("\\", "/")
        isExist = os.path.exists(fileDir)
        if not isExist:
            messages.warning(request, "No failed uploads.")
        else:
            filePathList = os.listdir(fileDir)
            filePathList = sorted(filePathList, key=lambda x: GenericLibraries().get_timestamp(x),reverse=True)
            if not filePathList:
                messages.warning(request, "No failed uploads.")
            filePaths    = Paginator(filePathList, 10)
            context.update({ 'filenames': filePaths })

        if request.method == 'POST':
            filePathList = os.listdir(fileDir)
            if request.POST.get('searchFile') and request.POST.get('searchFileDate'):
                fileDate = request.POST['searchFileDate'].split("-")
                fileYear = fileDate[0]
                fileMonth = fileDate[1]
                fileDay = fileDate[2]
                fileDate = fileYear+fileMonth+fileDay
                filePathList = [d for d in filePathList if fileDate in d]
                filePathList = [d for d in filePathList if request.POST['searchFile'].lower() in d.lower()]
            elif request.POST.get('searchFile'):
                filePathList = [d for d in filePathList if request.POST['searchFile'].lower() in d.lower()]
            elif request.POST.get('searchFileDate'):
                fileDate = request.POST['searchFileDate'].split("-")
                fileYear = fileDate[0]
                fileMonth = fileDate[1]
                fileDay = fileDate[2]
                fileDate = fileYear+fileMonth+fileDay
                filePathList = [d for d in filePathList if fileDate in d]
            filePathList = sorted(filePathList, key=lambda x: GenericLibraries().get_timestamp(x),reverse=True)
            
            if not filePathList:
                messages.warning(request, "No file found.")
            filePaths    = Paginator(filePathList, 10)
            for file in request.POST.lists():
                context['filterFile'] = request.POST['searchFile']
                context['filterFileDate'] = request.POST['searchFileDate']
                context['filenames'] = filePaths
        return render(request, 'Menu/FailedUploadList.html', context)
    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/member-list')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS,redirect_field_name=None)
def UploadHistoryView(request,companySalt):
    try:
        decodeCompanySalt = base64.b64decode(companySalt)
        decodeASCII = decodeCompanySalt.decode('UTF-8')
        getCompanyID = decodeASCII.split('_')
        companyID = getCompanyID[1]
        getUserCorporateObj = CorporateProfile.objects.get(id=companyID)
        context = {
            'company': getUserCorporateObj,
            'host_address' : settings.HOST_ADDRESS,
            'companySalt'  : GenericLibraries().saltEncode(getUserCorporateObj)
        }
        fileDir = os.path.join(settings.BASE_DIR,"media\\" + getUserCorporateObj.company_name + "\\Upload").replace("\\",'/')
        isExist = os.path.exists(fileDir)
        if not isExist:
            messages.warning(request,"No file uploaded.")
        else:
            filePathList = os.listdir(fileDir)
            filePathList = [fn for fn in filePathList if not fn.endswith('.pdf')]
            filePathList = sorted(filePathList, key=lambda x: GenericLibraries().get_timestamp(x),reverse=True)
            if not filePathList:
                messages.warning(request,'No file uploaded.')
            filePaths = Paginator(filePathList,10)
            context.update({ 'filenames':filePaths })

        if request.method == 'POST':
            if request.POST.get('searchFile') and request.POST.get('searchFileDate'):
                fileDate = request.POST['searchFileDate'].split("-")
                fileYear = fileDate[0]
                fileMonth = fileDate[1]
                fileDay = fileDate[2]
                fileDate = fileYear+fileMonth+fileDay
                filePathList = [d for d in filePathList if fileDate in d]
                filePathList = [d for d in filePathList if request.POST['searchFile'].lower() in d.lower()]
            elif request.POST.get('searchFile'):
                filePathList = [d for d in filePathList if request.POST['searchFile'].lower() in d.lower()]
            elif request.POST.get('searchFileDate'):
                fileDate = request.POST['searchFileDate'].split("-")
                fileYear = fileDate[0]
                fileMonth = fileDate[1]
                fileDay = fileDate[2]
                fileDate = fileYear+fileMonth+fileDay
                filePathList = [d for d in filePathList if fileDate in d]
            filePathList = sorted(filePathList, key=lambda x: GenericLibraries().get_timestamp(x),reverse=True)
            
            if not filePathList:
                messages.warning(request,'No file uploaded.')
            filePaths = Paginator(filePathList,10)
            for file in request.POST.lists():
                context['filterFile'] = request.POST['searchFile']
                context['filterFileDate'] = request.POST['searchFileDate']
                context['filenames'] = filePaths
        return render(request,'Menu/UploadHistoryList.html',context)
    except Exception as e:
        messages.error(request,str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/member-list')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def UploadMemberSpreadsheet(request):
    try:
        getUserCorporateObj = CorporateProfile.objects.get(user__id=request.user.id)
        keyFields  = [
                        {'employment_no':False}, 
                        {'name':True}, 
                        {'email_address':True}, 
                        {'mobile_no':True}, 
                        {'nationality':True}, 
                        {'mykad':True}, 
                        {'Death':False}, 
                        {'Disability':False}, 
                        {'Critical Illness':False}, 
                        {'Accident':False}, 
                        {'Medical':False}
                    ]

        try:
            fileUpload = request.FILES.getlist('member-spreadsheet')
            fileFolder = os.path.join(settings.BASE_DIR, "media/" + getUserCorporateObj.company_name + "/Upload")
            count      = 1
            isExist    = os.path.exists(fileFolder)
            if not isExist:
                os.mkdir(fileFolder)
            for file in os.listdir(fileFolder):
                count +=1
                
            currentDateTime = datetime.datetime.now()
            splitData = request.POST['member-form'].split('|')
            image_64_decode = base64.b64decode(splitData[1].split(',')[1])
            filenameUpload = splitData[0]
            filenameUpload = os.path.splitext(filenameUpload)[0]
            filename   = filenameUpload + '_upload_' + currentDateTime.strftime("%Y%m%d") + '_' + currentDateTime.strftime("%H%M") + '_' + str(count)
            fileExt    = '.xlsx'
            open(settings.MEDIA_ROOT + '/' + getUserCorporateObj.company_name + '/Upload/' + filename + fileExt, 'wb').write(image_64_decode)
            filePath   = fileFolder + '/' + filename + fileExt

            try:
                df         = pd.read_excel(filePath, sheet_name="Upload Members", engine='openpyxl')
                cleandf    = df.where(pd.notnull(df), None)
                toDict     = cleandf.to_dict(orient='records')
                rowCount   = 0
                batch_no   = 'BN' + str(datetime.date.today().year) + str(datetime.date.today().month) + '-' + str(get_next_value(sequence_name="registration/" + str(datetime.date.today().year) + str(datetime.date.today().month))).zfill(3)
            except (Exception, FileNotFoundError) as e:
                messages.error(request, "Please upload the excel file with the template provided.")
                logger.error(str(e), extra={'username':request.user.id})
                return HttpResponse(status=400)
                
            failToUploadIndividual = []
            headerExcel = []
            country_code = "+60"

            # Same file name
            existing_uploads = UploadMemberFilenames.objects.filter(corporate=getUserCorporateObj)
            if existing_uploads:
                for ext in existing_uploads:
                    if ext.original.lower() == fileUpload[0].name.replace(fileExt, "").lower():
                        messages.error(request, settings.SAME_FILENAME_MESSAGE)
                        responseData = {
                            'response_status' : 'Error',
                            'response_message': 'Upload Fail!'
                        }
                        return HttpResponse(json.dumps(responseData), content_type='application/json', status=400)

            new_upload = UploadMemberFilenames(
                corporate   = getUserCorporateObj,
                original    = fileUpload[0].name.replace(fileExt, ""),
                renamed     = filename
            )
            new_upload.save()
        except (Exception,FileNotFoundError) as e:
            messages.error(request, "No such file or directory.")
            logger.error(str(e),extra={'username':request.user.id})
            return redirect(settings.HOST_ADDRESS+'/member-list')

        deartimeDB          = DearTimeDbConn()
        isConnected         = deartimeDB.connect()
        if not isConnected:
            messages.error(request, settings.CONNECTION_LOST_MESSAGE)
            return redirect(settings.HOST_ADDRESS+'/member-list')
        productKeyFields = [prd.product_name for prd in Product.objects.filter(is_active=True)]
        for individual in toDict:
            dataDict = {}
            if rowCount > 2:
                hasError = False

                # Change dictionary key name
                dataCount = 0
                colCount = 0
                for key, data in individual.items():
                    if colCount != 6 and colCount!=7 and colCount!=8:
                        dataDict[list(keyFields[dataCount])[0]] = data
                        dataCount+=1
                    colCount+=1
                
                # Change None to 0
                for prd in productKeyFields:
                    if dataDict[prd] == None:
                        dataDict[prd] = 0

                # Rounded coverages for certain validation and insertion
                try:
                    roundedDict = {}
                    for prd in productKeyFields:
                        roundedDict[prd] = round(int(dataDict[prd]) / 1000) * 1000

                    checkMemberConditions = CheckMemberConditions()
                    coverage_dict = {}
                    for prd in productKeyFields:
                        coverage_dict[prd] = int(dataDict[prd])
                    
                    # Check coverage limit
                    hasError = checkMemberConditions.check_coverage_limit(coverage_dict, 'single')
                    if hasError:
                        if dataDict.get('reason') is None:
                            hasError = True
                            dataDict['reason'] = [settings.COVERAGE_LIMIT_MESSAGE]                   
                        else:
                            hasError = True
                            dataDict['reason'].append(settings.COVERAGE_LIMIT_MESSAGE)

                    # Validate atleast one product coverage need to be filled
                    hasCoverage = False
                    for prd in productKeyFields:
                        if dataDict[prd] != None and int(dataDict[prd]) != 0:
                            hasCoverage = True

                    if hasCoverage == False:
                        if dataDict.get('reason') is None:
                            hasError = True
                            dataDict['reason'] = [settings.COVERAGE_MANDATORY_MESSAGE]
                        else:
                            hasError = True
                            dataDict['reason'].append(settings.COVERAGE_MANDATORY_MESSAGE)
                    
                    # Accident coverage must be less than death coverage
                    if dataDict['Accident'] != None:
                        if dataDict['Death'] == None or (int(dataDict['Death']) < int(dataDict['Accident'])):
                            if dataDict.get('reason') is None:
                                hasError = True
                                dataDict['reason'] = [settings.ACCIDENT_COVERAGE_MESSAGE]
                            else:
                                hasError = True
                                dataDict['reason'].append(settings.ACCIDENT_COVERAGE_MESSAGE)
    
                    # Medical coverage must follow plans
                    getMedicalObj   = GetMedicalPlans()
                    medical_deductibles = getMedicalObj.getMedical(deartimeDB)
                    if dataDict['Medical'] and int(dataDict['Medical']) != 0:
                        if not int(dataDict['Medical']) in medical_deductibles:
                            if dataDict.get('reason') is None:
                                hasError = True
                                dataDict['reason'] = [settings.MEDICAL_COVERAGE_MESSAGE]
                            else:
                                hasError = True
                                dataDict['reason'].append(settings.MEDICAL_COVERAGE_MESSAGE)  

                    # Check minimum coverage
                    for k, v in roundedDict.items():
                        if k != 'Medical':
                            getMinimumCoverage = deartimeDB.exec_SQL('getPremiumRate', {'PRODUCT_NAME': k}, 'fetchone')
                            options_dict = json.loads(getMinimumCoverage['dset'][0])
                            minimum_coverage = options_dict['min_coverage']
                            if v != 0 and v < minimum_coverage:
                                if dataDict.get('reason') is None:
                                    hasError = True
                                    dataDict['reason'] = [settings.MINIMUM_COVERAGE_MESSAGE.format(PRODUCT=k)]                   
                                else:
                                    hasError = True
                                    dataDict['reason'].append(settings.MINIMUM_COVERAGE_MESSAGE.format(PRODUCT=k))       

                except:
                    if dataDict.get('reason') is None:
                        hasError = True
                        dataDict['reason'] = [settings.COVERAGE_SHOULD_BE_INT]           
                    else:
                        hasError = True
                        dataDict['reason'].append(settings.COVERAGE_SHOULD_BE_INT)             

                # Validate mandatory fields if empty
                dataCount = 0
                for field in keyFields:
                    if list(field.values())[0] and dataDict[list(field)[0]] is None:
                        if dataDict.get('reason') is None:
                            hasError = True
                            dataDict['reason'] = [settings.MANDATORY_MESSAGE]
                            break
                        else:
                            hasError = True
                            dataDict['reason'].append(settings.MANDATORY_MESSAGE)
                            break
                    dataCount+=1      

                #Formatting mobile number
                formatted_mobile_number = re.sub('[^0-9+]+', '', str(dataDict['mobile_no']))
                if country_code in formatted_mobile_number:
                    formatted_mobile_number = formatted_mobile_number[2:]

                try:
                    if formatted_mobile_number[0] != '0':
                        formatted_mobile_number = formatted_mobile_number[1:]
                except:
                    if dataDict.get('reason') is None:
                        hasError = True
                        dataDict['reason'] = [settings.MOBILE_SHOULD_BE_INT]           
                    else:
                        hasError = True
                        dataDict['reason'].append(settings.MOBILE_SHOULD_BE_INT)       

                #formatted nric
                formatted_mykad = re.sub('[^0-9]+', '', str(dataDict['mykad']))
                
                try:
                    nric = formatted_mykad
                    year = nric[0:2]
                    month = nric[2:4]
                    day = nric[4:6]
                    today = datetime.datetime.today()
                    cutoff = today.year - 2000
                    if int(year)>int(cutoff):
                        year='19'+year
                    else:
                        year='20'+year
                    dob = year+'-'+month+'-'+day
                    try:
                        dob = datetime.datetime.strptime(dob, "%Y-%m-%d")
                        
                        if dob > today:
                            if dataDict.get('reason') is None:
                                hasError = True
                                dataDict['reason'] = [settings.NRIC_DOB_NOTVALID_MESSAGE]
                            else:
                                hasError = True
                                dataDict['reason'].append(settings.NRIC_DOB_NOTVALID_MESSAGE)
                            
                    except:
                        if dataDict.get('reason') is None:
                            hasError = True
                            dataDict['reason'] = [settings.NRIC_DOB_NOTVALID_MESSAGE]
                        else:
                            hasError = True
                            dataDict['reason'].append(settings.NRIC_DOB_NOTVALID_MESSAGE)

                    last_digit = nric[10:14]
                    if int(last_digit) % 2 != 0:
                        genderData = 'male'
                    else:
                        genderData = 'female'

                    getAge = AgeCalculator().calculate_age(dob)
                    age_limit_dict = {
                        'Death'             : 65,
                        'Disability'        : 65,
                        'Critical Illness'  : 60,
                        'Accident'          : 65,
                        'Medical'           : 55,
                    }

                    try:
                        if getAge != -1:
                            for prd2 in productKeyFields:
                                if int(dataDict[prd2]) != 0:
                                    if getAge > age_limit_dict[prd2.title()]:
                                        if dataDict.get('reason') is None:
                                            hasError = True
                                            dataDict['reason'] = [settings.AGE_LIMIT_MESSAGE.format(PRODUCT=prd2.title())]
                                        else:
                                            hasError = True
                                            dataDict['reason'].append(settings.AGE_LIMIT_MESSAGE.format(PRODUCT=prd2.title()))
                    except:
                        if dataDict.get('reason') is None:
                            hasError = True
                            dataDict['reason'] = [settings.COVERAGE_SHOULD_BE_INT]           
                        else:
                            hasError = True
                            dataDict['reason'].append(settings.COVERAGE_SHOULD_BE_INT)     
                except:
                    if dataDict.get('reason') is None:
                        hasError = True
                        dataDict['reason'] = [settings.NRIC_DOB_NOTVALID_MESSAGE]           
                    else:
                        hasError = True
                        dataDict['reason'].append(settings.NRIC_DOB_NOTVALID_MESSAGE)

                if len(nric) != 12:
                    if dataDict.get('reason') is None:
                        hasError = True
                        dataDict['reason'] = [settings.NRIC_DOB_NOTVALID_MESSAGE]
                    else:
                        hasError = True
                        if settings.NRIC_DOB_NOTVALID_MESSAGE not in dataDict['reason']:
                            dataDict['reason'].append(settings.NRIC_DOB_NOTVALID_MESSAGE)

                # Check existing member in CPFO
                checkCPFOExistMember = GenericLibraries().checkCPFOExistMember(dataDict['email_address'], formatted_mobile_number, formatted_mykad, getUserCorporateObj)

                if not checkCPFOExistMember:
                    if dataDict.get('reason') is None:
                        hasError = True
                        dataDict['reason'] = [settings.EMAIL_MOBILE_EXISTING_MESSAGE]
                    else:
                        hasError = True
                        dataDict['reason'].append(settings.EMAIL_MOBILE_EXISTING_MESSAGE)

                if not hasError:
                    # Check existing member in Deartime DB
                    if dataDict['employment_no'] is None:
                        dataDict['employment_no'] = '-'
                    # Save to Deartime DB
                    getExistEmailOrMobile = deartimeDB.exec_SQL('validateExistingEmailMobile', {'MOBILE': formatted_mobile_number, 'EMAIL' : dataDict['email_address'], 'NRIC': formatted_mykad}, 'fetchone')

                    if getExistEmailOrMobile['dset']:
                        getExistMember        = deartimeDB.exec_SQL('validateMember', {'MOBILE': formatted_mobile_number, 'EMAIL' : dataDict['email_address'], 'NRIC' : formatted_mykad}, 'fetchone')
                        
                        if not getExistMember['dset']:
                            if dataDict.get('reason') is None:
                                hasError = True
                                dataDict['reason'] = [settings.EMAIL_MOBILE_EXISTING_MESSAGE]
                            else:
                                hasError = True
                                dataDict['reason'].append(settings.EMAIL_MOBILE_EXISTING_MESSAGE)
                        else:
                            memberID  = getExistMember['dset'][0]

                            #Check prospect covered under sponsored insurance
                            getSponsoredInsurance = deartimeDB.exec_SQL('getSponsoredInsurance', {'USER_ID': memberID}, 'fetchone')

                            if getSponsoredInsurance['dset']:
                                if int(last_digit) % 2 != 0:
                                    gender = 'him'
                                    gender2 = 'his'
                                else:
                                    gender = 'her'
                                    gender2 = 'her'
                                if dataDict.get('reason') is None:
                                    hasError = True
                                    dataDict['reason'] = [settings.UNDER_SPONSORED_INSURANCE.format(INSURED_NAME=getSponsoredInsurance['dset'][0], NEXT_RENEWAL_DATE=datetime.datetime.strftime(getSponsoredInsurance['dset'][1], '%d %B %Y'),GENDER=gender, GENDER2=gender2)] 
                                else:
                                    hasError = True
                                    dataDict['reason'] = [settings.UNDER_SPONSORED_INSURANCE.format(INSURED_NAME=getSponsoredInsurance['dset'][0], NEXT_RENEWAL_DATE=datetime.datetime.strftime(getSponsoredInsurance['dset'][1], '%d %B %Y'),GENDER=gender, GENDER2=gender2)]

                            if not hasError:
                                getIndividualID  = deartimeDB.exec_SQL('getIndividual', {'USER_ID': getExistMember['dset'][0]}, 'fetchone')

                                # Check existing medical coverage
                                if dataDict['Medical']:
                                    checkActiveMedical = GenericLibraries().checkActiveMedical(deartimeDB, getIndividualID, dataDict['Medical'])
                                    if checkActiveMedical:
                                        if dataDict.get('reason') is None:
                                            hasError = True
                                            dataDict['reason'] = [settings.MEDICAL_EXISTING_MESSAGE]
                                        else:
                                            hasError = True
                                            dataDict['reason'].append(settings.MEDICAL_EXISTING_MESSAGE)

                    if not hasError:
                        # Coverages rounding
                        for prd in productKeyFields:
                            if prd != 'Medical':
                                dataDict[prd] = int(GenericLibraries().round_half_up(dataDict[prd]/1000) * 1000)

                        checkExistPackageObj = CheckUniquePackage()
                        productToCheck = {}
                        for prd in productKeyFields:
                            productToCheck[prd] = int(dataDict[prd])                            
                        existPackage = checkExistPackageObj.check(productToCheck, getUserCorporateObj.user_id)
                        if not existPackage:
                            createPackage = Package(
                                package_name = 'Customized-' + str(get_next_value(sequence_name=getUserCorporateObj.company_name+"_package/")).zfill(3),
                                created_by   = request.user
                            )
                            createPackage.save()

                            productData = []
                            productDict = {}
                            for prd in productKeyFields:
                                getProductObj = Product.objects.filter(product_name__icontains=prd).first()
                                if getProductObj:
                                    if dataDict[prd] == None:
                                        dataDict[prd] = 0
                                    productDict.update({
                                        'package'        : createPackage,
                                        'product'        : getProductObj,
                                        'coverage_amount': dataDict[prd],
                                        'created_by'     : request.user
                                    })
                                    productData.append(
                                        PackageProductMapping(**productDict)
                                    )
                                
                            if productData:
                                with transaction.atomic():
                                    PackageProductMapping.objects.bulk_create(productData)
                            
                            dataDict.update({
                                'corporate_id'     : getUserCorporateObj.id,
                                'submitted'        : False,
                                'package'          : createPackage,
                                'status'           : 'Pending',
                                'batch_no'         : batch_no,
                                'mobile_no'        : formatted_mobile_number,
                                'dob'              : dob,
                                'gender'           : genderData,
                                'mykad'            : formatted_mykad
                                
                            })
                        else:
                            dataDict.update({
                                'corporate_id'     : getUserCorporateObj.id,
                                'submitted'        : False,
                                'package_id'       : existPackage.id,
                                'status'           : 'Pending',
                                'batch_no'         : batch_no,
                                'mobile_no'        : formatted_mobile_number,
                                'dob'              : dob,
                                'gender'           : genderData,
                                'mykad'            : formatted_mykad
                            })

                        dataDict.pop('Death')
                        dataDict.pop('Disability')
                        dataDict.pop('Critical Illness')
                        dataDict.pop('Accident')
                        dataDict.pop('Medical')

                        if getExistEmailOrMobile['dset']:
                            if getExistMember['dset']:
                                dataDict['deartime_memberid'] = memberID
                                if getExistMember['dset'][1]:
                                    dataDict['is_existing'] = True        

                            #Check prospect on sponsored insurance waiting list
                            getSponsoredInsuranceWaitingList = GenericLibraries().checkSIWaitingList(deartimeDB, memberID)

                            if getSponsoredInsuranceWaitingList:
                                dataDict['si_waitinglist'] = True
                                dataDict['siwaiting_email'] = True

                                #update SI waiting list to DT database
                                updateSIConfirm = deartimeDB.exec_SQL('updateSponsoredInsuranceConfirm', {'USER_ID': memberID}, 'update')

                        mappedData = Member(**dataDict)
                        
                        if mappedData:
                            with transaction.atomic():
                                mappedData.save()
                                getMember = mappedData

                        calculator = PremiumCalculator()
                        memberProductData   = []
                        if getMember:
                            getNewMappings      = PackageProductMapping.objects.filter(package=getMember.package)
                            if getNewMappings:
                                for nmpg in getNewMappings:
                                    memberProductDict = {
                                        'member_id'             : getMember.id,
                                        'product_id'            : nmpg.product_id,
                                        'coverage_amount'       : nmpg.coverage_amount
                                    }
                                    memberProductData.append(
                                        MemberProductMapping(**memberProductDict)
                                    )
                                    if memberProductData:
                                        with transaction.atomic():
                                            getNewMemberProductMappings = MemberProductMapping.objects.bulk_create(memberProductData)
                                            memberProductData = []
                                
                            quotationDict = calculator.calculate_quotation(deartimeDB, getMember, getUserCorporateObj.payment_mode)
                            getMember.quotation_premium = quotationDict
                            getMember.tentative_premium = quotationDict
                            getMember.save()
                    else:
                        failToUploadIndividual.append(dataDict)
                else:
                    failToUploadIndividual.append(dataDict)
            else:
                if rowCount == 0:
                    for key, data in individual.items():
                        headerExcel.append(data)
            rowCount+=1

        deartimeDB.close()
        if failToUploadIndividual:
            try:
                # Create a new workbook and add a worksheet.
                filePathx       = os.path.join(settings.BASE_DIR, "media/" + getUserCorporateObj.company_name + "/FailUploadMember")
                isExist         = os.path.exists(filePathx)
                if not isExist:
                    os.mkdir(filePathx)

                workBookName    = filePathx + "/" + filename + '.xlsx'
                workbook        = xlsxwriter.Workbook(workBookName)
                worksheet       = workbook.add_worksheet() # Default to Sheet1 if no name is defined

                headerColNum = 1
                # Add a bold format to use to highlight cells.
                # Add text wrap for 'reason' column for duplicate reasons
                bold         = workbook.add_format({'bold': True})
                text_wrap    = workbook.add_format({'text_wrap': True})
                for head in headerExcel:
                    if head!='Gender' and head!='Passport No' and head!='Date of Birth':
                        worksheet.write(0, headerColNum, head, bold)
                        headerColNum += 1
                worksheet.write(0, headerColNum, 'Reason', bold)

                # Start from the second cell. Rows and columns are zero indexed.
                row = 1
                # Iterate over the data and write it out row by row.
                for dataItems in failToUploadIndividual:
                    col = 0
                    for dataKey, dataValue in dataItems.items():
                        if dataKey == 'reason':
                            col+=1
                            worksheet.write(row, col, "\n".join(dataValue), text_wrap)
                        elif dataKey == 'dob':
                            if col == 0:
                                worksheet.write(row, col, col+1)
                            col+=1
                            dateFormat = workbook.add_format({'num_format': 'dd/mm/yyyy'})
                            worksheet.write(row, col, dataValue, dateFormat)
                        else:
                            if col == 0:
                                worksheet.write(row, col, row)
                            col+=1
                            worksheet.write(row, col, dataValue)
                            # col += 1
                    row += 1
                workbook.close()

                # Remove temporary excel file
                #default_storage.delete(filePath)

            except (Exception,FileNotFoundError) as e:
                messages.error(request, "No such file or directory.")
                logger.error(str(e),extra={'username':request.user.id})
                return redirect(settings.HOST_ADDRESS+'/member-list')
           
        if len(failToUploadIndividual) != (len(toDict) - 3):
            messages.success(request, 'Successfully uploaded ' + str(len(toDict) - 3 - len(failToUploadIndividual)) + 
                            ' member(s) with ' + str(len(failToUploadIndividual)) + ' failed uploads')
        else:
            messages.error(request, 'All members have failed to upload!')

        responseData = {
            'response_status' : 'Success',
            'response_message': 'Successfully uploaded!'
        }
        return HttpResponse(json.dumps(responseData), content_type='application/json', status=200)

    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
        responseData = {
            'response_status' : 'Error',
            'response_message': 'Upload Fail!'
        }
        return HttpResponse(json.dumps(responseData), content_type='application/json', status=400)

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def DeleteMember(request):
    try:
        deartimeDB = DearTimeDbConn()
        isConnected = deartimeDB.connect()
        if not isConnected:
            messages.error(request, settings.CONNECTION_LOST_MESSAGE)
            return redirect(settings.HOST_ADDRESS+'/member-list')
        with transaction.atomic():
            getDeleteMember = Member.objects.get(id=request.POST['terminatedMember'])
            getDeleteMember.void = True
            getDeleteMember.status = 'Terminated'
            getDeleteMember.rejected_reason = 'Corporate Rescinded The Offer'
            getDeleteMember.save()
        corporateObj = CorporateProfile.objects.get(id=getDeleteMember.corporate_id)
        getPackageProductMapping = PackageProductMapping.objects.filter(created_by_id=request.user.id, package_id=getDeleteMember.package_id)
        
        for prodMap in getPackageProductMapping:
            getExistingMemberPrdMap = MemberProductMapping.objects.get(member_id=getDeleteMember.id, product_id=prodMap.product_id, is_terminated=False, is_renewal=False)
            getExistingMemberPrdMap.is_terminated = True
            getExistingMemberPrdMap.updated_datetime = datetime.datetime.now()
            getExistingMemberPrdMap.save()
            if getExistingMemberPrdMap.deartime_coverageid:
                getIndividual = deartimeDB.exec_SQL('getIndividual', {'USER_ID': getDeleteMember.deartime_memberid}, 'fetchone')    
                getProductID = deartimeDB.exec_SQL('getProductObj', {'PRD_NAME': prodMap.product.product_name.title()}, 'fetchone')
                deartimeDB.exec_SQL('updateSingleTerminateCoverage', {'OWNER_ID': getIndividual['dset'][0], 'PAYER_ID': corporateObj.deartime_payerid, 'PRODUCT_ID': getProductID['dset'][0]}, 'update')
          
        deartimeDB.close()
        messages.success(request, 'Successfully terminated!')
        return redirect(settings.HOST_ADDRESS+'/member-list')
    except Exception as e:
        context = {
            'host_address': settings.HOST_ADDRESS
        }
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return render(request, 'Menu/MemberList.html', context)

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def ReofferMember(request):
    try:
        deartimeDB = DearTimeDbConn()
        isConnected = deartimeDB.connect()
        if not isConnected:
            messages.error(request, settings.CONNECTION_LOST_MESSAGE)
            return redirect(settings.HOST_ADDRESS+'/member-list')
        with transaction.atomic():
            isMedicalSurvey = False
            getExistingMember = None
            for info in request.POST:
                if info == 'unterminatedMember':
                    splitmember = request.POST.get('unterminateMember').split(',')
                    for member in splitmember:
                        getExistingMember = Member.objects.get(id=member)
                        existingMemberStatus = 'Terminated'
                        getExistingMember.void = False
                        getExistingMember.paid = False
                        getExistingMember.medical_survey = False
                        getExistingMember.status = 'Pending'
                        getExistingMember.save()

                elif info == 'unrejectMember':
                    splitmember = request.POST.get('reofferMember').split(',')
                    for member in splitmember:
                        getExistingMember = Member.objects.get(id=member)
                        if getExistingMember.rejected_reason.lower() == 'rejected by medical survey':
                            isMedicalSurvey = True
                        existingMemberStatus = 'Reject'
                        getExistingMember.rejected = False
                        getExistingMember.status = 'Pending'
                        getExistingMember.rejected_reason = None
                        getExistingMember.save()

            for member in splitmember:
                getExistingMember = Member.objects.get(id=member)
                if not isMedicalSurvey:
                    getMemberProductMapping = MemberProductMapping.objects.filter(member_id = getExistingMember.id, is_terminated = False)
                    if getMemberProductMapping:
                        for mpm in getMemberProductMapping:
                            mpm.is_terminated = True
                            mpm.save()

                    getNewMappings  = PackageProductMapping.objects.filter(package=getExistingMember.package)
                    if getNewMappings:
                        for nmpg in getNewMappings:
                            new_member_product_mapping = MemberProductMapping(
                                coverage_amount     = nmpg.coverage_amount,
                                member_id           = getExistingMember.id,
                                product             = nmpg.product
                            )
                            new_member_product_mapping.save()
                        
            getUserCorporateObj = CorporateProfile.objects.get(user__id=request.user.id)
            getMemberObj = Member.objects.get(id=getExistingMember.id)
            calculator = PremiumCalculator()
            quotationDict = calculator.calculate_quotation(deartimeDB, getMemberObj, getUserCorporateObj.payment_mode)
            getMemberObj.quotation_premium = quotationDict
            getMemberObj.tentative_premium = quotationDict
            getMemberObj.save()

        deartimeDB.close()
        messages.success(request, 'Successfully reinstate member!')
        return redirect(settings.HOST_ADDRESS+'/member-list')
    except Exception as e:
        context = {
            'host_address': settings.HOST_ADDRESS
        }
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return render(request, 'Menu/MemberList.html', context)

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def GenerateInvoicePDF(request):
    try:
        # Save to Deartime DB
        deartimeDB  = DearTimeDbConn()
        isConnected = deartimeDB.connect()
        if not isConnected:
            messages.error(request, settings.CONNECTION_LOST_MESSAGE)
            return redirect(settings.HOST_ADDRESS+'/member-list')
        else:
            getUserCorporateObj   = CorporateProfile.objects.get(user_id=request.user.id)
            getUserCorporateRefNo = deartimeDB.exec_SQL('getUserRefNo', {'USER_ID': getUserCorporateObj.deartime_payerid}, 'fetchone')
            generateInvoiceNo = 'N' + str(datetime.date.today().year) + str(datetime.date.today().month).zfill(2) + '/' + str(get_next_value(sequence_name="invoice")).zfill(5) + str(settings.INVOICE_GENERATION_REFERENCE)
            total         = 0
            true_total    = 0

            memberList = []
            selectedMember = {}
            calculator = PremiumCalculator()
            for key, value in request.POST.lists():
                if "paymentmember" in key:
                    memberID = int(key.removeprefix('paymentmember').removesuffix("_POST"))
                    getMember = Member.objects.get(id=memberID)
                    #getMemberProductMapping = MemberProductMapping.objects.filter(member_id=getMember.id, is_terminated=False).exclude(deartime_coverageid__isnull=True)
                    getMemberProductMapping = MemberProductMapping.objects.filter(member_id=getMember.id, is_terminated=False, is_renewal= getMember.renew)
                    if 'renew' in request.POST:
                        memberIndividual       = deartimeDB.exec_SQL('getIndividual', {'USER_ID': getMember.deartime_memberid}, 'fetchone')
                        getRenewalMPM = MemberProductMapping.objects.filter(member_id=getMember.id, is_terminated=False, is_renewal=True)
                        if getRenewalMPM:
                            for memberProduct in getRenewalMPM:
                                updateCoverageDate = deartimeDB.exec_SQL('updateCoverageDateWithoutStatus', {'UPDATED_DATE': str(datetime.datetime.now()), 'COVERAGE_ID':memberProduct.deartime_coverageid}, 'update')
                                memberProduct.updated_datetime = datetime.datetime.now()
                                memberProduct.save()
                        else:
                            insertCoverage = GenericLibraries().insertCoverageToDTDB(deartimeDB, memberID, memberIndividual)
                    getMemberADPremium = PremiumAdjustment.objects.filter(member_id = getMember.id, generated_invoice = False, paid = False, void = False)
                    if not getMemberADPremium:
                        getMember.tentative_premium = calculator.calculate_premium(getMember.deartime_memberid, getUserCorporateObj.id, 'total', getMemberProductMapping, deartimeDB, 'renew' in request.POST)
                        getMember.true_premium = calculator.calculate_premium(getMember.deartime_memberid, getUserCorporateObj.id, 'total', getMemberProductMapping, deartimeDB, 'renew' in request.POST, None, True)
                        getMember.generated_invoice = True
                        getMember.save()
                        total += getMember.tentative_premium
                    else: 
                        for memberObj in getMemberADPremium:
                            getMember.tentative_premium = float(getMember.tentative_premium) + float(memberObj.ad_premium)
                            getMember.save()
                            memberObj.generated_invoice = True
                            memberObj.save()
                            total += float(memberObj.ad_premium)        
                        
                    true_total += getMember.true_premium
                    memberList.append(getMember)
                    selectedMember.update({ getMember.employment_no : getMember.id })

                    new_message_queue = MessagingQueue(
                        email_address = getMember.email_address,
                        module        = 'GenerateContractView'
                    )
                    new_message_queue.save()
                             
            totalPayables = "{:.2f}".format(total)
            totalPayablesWord = num2words(totalPayables, to='currency').title().replace('Euro', "")
            if float(totalPayables).is_integer():
                totalPayablesWord = totalPayablesWord.split(',')[0]
            totalPayablesWord += " Only"

            rowCount = 10
            paginator = Paginator(memberList, rowCount)
            
            if paginator.object_list:
                for member in paginator.object_list:
                    if (PremiumAdjustment.objects.filter(member_id=member.id, paid=False, generated_invoice=True, void=False).exists()):
                        getMemberADPremium = PremiumAdjustment.objects.get(member_id=member.id, paid=False, generated_invoice=True, void=False)
                        member.ad_premium = getMemberADPremium.ad_premium

            saveInvoice = Invoice (
                company          = getUserCorporateObj,
                invoice_no       = generateInvoiceNo,
                total_amount     = totalPayables,
                created_by       = request.user,
                status           = 'Pending Payment'
            )
            saveInvoice.save()
            if settings.ENVIRONMENT_INDICATOR != '':
                preferredCurrentDate = CurrentDate.objects.filter(corporate_id=getUserCorporateObj.id)
                if preferredCurrentDate:
                    currentObject = CurrentDate.objects.get(corporate_id=getUserCorporateObj.id)
                    saveInvoice.created_datetime = currentObject.current_datetime
                    saveInvoice.save()

            # Current date
            currentDate = GenericLibraries().currentDateTesting(getUserCorporateObj.id, True)

            file = GenericLibraries().render_to_pdf(getUserCorporateObj, getUserCorporateObj.company_name,  generateInvoiceNo, 'InvoiceAndPayment/Invoice.html',
                {
                    'pagesize'           : 'A4',
                    'image_url'          : settings.STATICFILES_DIRS[0].replace("\\","/") + '/portal/img/deartime-logo-inverted-color.png',
                    'company'            : getUserCorporateObj,
                    'tables'             : paginator,
                    'total_payables'     : totalPayables,
                    'total_payables_text': totalPayablesWord,
                    'payor_ref'          : getUserCorporateRefNo['dset'][0],
                    'invoice_no'         : generateInvoiceNo,
                    'invoice_date'       : datetime.datetime.strftime(Invoice.objects.get(company_id=getUserCorporateObj.id, invoice_no=generateInvoiceNo).created_datetime, '%d %B %Y'),
                    'flag'               : 'invoice'
                })
   
            for member in memberList:
                generateOrderNo = 'ORD' + str(datetime.date.today().year) + str(datetime.date.today().month) + '/' + str(get_next_value(sequence_name=getUserCorporateObj.company_name+"_order/")).zfill(5)
                saveOrder = Order(
                    invoice    = saveInvoice,
                    member     = member,
                    order_no   = generateOrderNo,
                    amount     = member.tentative_premium,
                    true_amount= member.true_premium,
                    created_by = request.user
                )
                saveOrder.save()

            getLatestOrderID        = deartimeDB.exec_SQL('selectMaxIDOrder', {}, 'fetchone')
            getLatestTransactionID  = deartimeDB.exec_SQL('selectMaxIDTransaction', {}, 'fetchone')
            nextLatestOrderID       = getLatestOrderID['dset'][0] + 1
            nextLatestTransactionID = getLatestTransactionID['dset'][0] + 1
            orderRefNo              = 'OR' + str(nextLatestOrderID).zfill(6)
            transactionRefNo        = 'TX' + str(nextLatestTransactionID).zfill(6)
            nextTryOn               = datetime.datetime.now() + datetime.timedelta(days=7)
            if 'renew' in request.POST:
                dataDictsOR             = (str(uuid.uuid4()), totalPayables, true_total, getUserCorporateObj.deartime_payerid, str(nextTryOn), str(datetime.datetime.now()), str(datetime.datetime.now()), 0, orderRefNo, str(datetime.datetime.now()), str(datetime.datetime.now()))
                getNewOrderID           = deartimeDB.exec_SQL('insertOrderRenewal', dataDictsOR, 'insert')
            else:
                dataDictsOR             = (str(uuid.uuid4()), totalPayables, true_total, getUserCorporateObj.deartime_payerid, str(nextTryOn), str(datetime.datetime.now()), str(datetime.datetime.now()), 0, orderRefNo, str(datetime.datetime.now()), str(datetime.datetime.now()))
                getNewOrderID           = deartimeDB.exec_SQL('insertOrder', dataDictsOR, 'insert')

            dataDictsTR             = (str(uuid.uuid4()), getNewOrderID['lastID'], 'manual', 'TRX'+(str(time())).split(".")[0], totalPayables, transactionRefNo, str(datetime.datetime.now()), str(datetime.datetime.now()), str(datetime.datetime.now()), 'FPX-B2B', getUserCorporateObj.company_name)
            getNewTransactionID     = deartimeDB.exec_SQL('insertTransaction', dataDictsTR, 'insert')
            for mem in memberList:
                getIndividual       = deartimeDB.exec_SQL('getIndividual', {'USER_ID': mem.deartime_memberid}, 'fetchone')
                getUnpaidCoverage   = deartimeDB.exec_SQL('getUnpaidCoverage', {'OWNER_ID': getIndividual['dset'][0], 'PAYER_ID': getUserCorporateObj.deartime_payerid}, 'fetchall')
                if getUnpaidCoverage:
                    for coverid in getUnpaidCoverage['dset']:
                        dataDictsCO         = (coverid[0], getNewOrderID['lastID'], str(datetime.datetime.now()), str(datetime.datetime.now()))
                        getNewCoverageOrderID = deartimeDB.exec_SQL('insertCoverageOrder', dataDictsCO, 'insert')
                        dataDictsCoverages  = {'CSD_INVOICE_DATE': str(currentDate), 'UPDATED_DATE':str(datetime.datetime.now()), 'COVERAGE_ID': coverid[0]}
                        updateCoveragesCSD  = deartimeDB.exec_SQL('updateCoveragesCSD', dataDictsCoverages, 'update')
                        if 'renew' in request.POST:
                            paymentDueDate = datetime.datetime.strptime(getUserCorporateObj.payment_due_date, '%Y-%m-%d')
                            if (currentDate.day < paymentDueDate.day) and (currentDate.month <=  paymentDueDate.month):
                                paymentDueDate = paymentDueDate + relativedelta(years=1)
                            dataDictsNDD        = {'NDD_PAYMENT_DUE_DATE':str(paymentDueDate), 'UPDATED_DATE':str(datetime.datetime.now()), 'COVERAGE_ID': coverid[0]}
                            updateCoveragesNDD  = deartimeDB.exec_SQL('updateMemberCoverageNDD', dataDictsNDD, 'update')
                        else:
                            dataDictsNDD        = {'NDD_PAYMENT_DUE_DATE':getUserCorporateObj.payment_due_date, 'UPDATED_DATE':str(datetime.datetime.now()), 'COVERAGE_ID': coverid[0]}
                            updateCoveragesNDD  = deartimeDB.exec_SQL('updateMemberCoverageNDD', dataDictsNDD, 'update')
                
            saveInvoice.deartime_orderid = getNewOrderID['lastID']
            saveInvoice.save()

            for mem2 in memberList:
                checkMemberADPremium = PremiumAdjustment.objects.filter(member_id = getMember.id, generated_invoice = True, paid = False, void = False)
                if not checkMemberADPremium:
                    mem2.status = 'P.Invoice'
                    mem2.save()

            strToHash  = settings.PROD_SENANGPAY_SECRET_KEY + saveInvoice.invoice_no.replace("/", "-") + str(saveInvoice.total_amount) + saveInvoice.invoice_no.replace("/", "-")
            sha256hash = hmac.new(bytes(settings.PROD_SENANGPAY_SECRET_KEY, 'UTF-8'), bytes(strToHash, 'UTF-8'), hashlib.sha256)
            saveInvoice.hash_value = sha256hash.hexdigest()
            saveInvoice.save()

            context = {
                'file'           : file,
                'path'           : request._current_scheme_host,
                'company'        : getUserCorporateObj,
                'invoiceNo'      : generateInvoiceNo,
                'selectedMembers': selectedMember,
                'flag'           : 'invoice',
                'invoiceObj'     : saveInvoice,
                'senangpay_url'  : settings.PROD_SENANGPAY_URL + settings.PROD_SENANGPAY_MERCHANT_KEY,
                'invoice_no'     : saveInvoice.invoice_no.replace("/", "-"),
                'host_address' : settings.HOST_ADDRESS
            }

            deartimeDB.close()
            return render(request, 'InvoiceAndPayment/InvoiceView.html', context)
    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/member-list')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def InvoicePayment(request):
    try:
        getUserCorporateObj = CorporateProfile.objects.get(user_id=request.user.id)
        getInvoiceObj = Invoice.objects.get(company=getUserCorporateObj, invoice_no=request.POST['invoiceNo'])
        getInvoiceObj.status = 'Payment In Progress'
        getInvoiceObj.click_datetime = datetime.datetime.now()
        getInvoiceObj.save()

        return JsonResponse({'message': 'success'})
    except Exception as ex:
        logger.error(str(ex),extra={'username':request.user.id})
        return JsonResponse({'message': 'error'})

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def ViewInvoiceView(request, invoiceSalt, type):
    decodeInvoiceSalt = base64.b64decode(invoiceSalt)
    decodeASCII = decodeInvoiceSalt.decode('UTF-8')
    getinvoiceID = decodeASCII.split('_')
    invoiceID = getinvoiceID[1]
    getCompanyObj   = CorporateProfile.objects.get(user_id=request.user.id)
    getInvoice      = Invoice.objects.get(id=invoiceID)
    if type == 'invoice':
        file = getCompanyObj.company_name + "/Invoice/" + getInvoice.invoice_no.replace("/", "_") + ".pdf"
        if not getInvoice.hash_value:
            strToHash  = settings.PROD_SENANGPAY_SECRET_KEY + getInvoice.invoice_no.replace("/", "-") + str(getInvoice.total_amount) + getInvoice.invoice_no.replace("/", "-")
            sha256hash = hmac.new(bytes(settings.PROD_SENANGPAY_SECRET_KEY, 'UTF-8'), bytes(strToHash, 'UTF-8'), hashlib.sha256)
            getInvoice.hash_value = sha256hash.hexdigest()
            getInvoice.save()
    elif type == 'receipt':
        file = getCompanyObj.company_name + "/Receipt/" + getInvoice.receipt_no.replace("/", "_") + ".pdf"

    context = {
        'file'         : file,
        'path'         : request._current_scheme_host,
        'company'      : getCompanyObj,
        'invoiceNo'    : getInvoice.invoice_no if type == 'invoice' else getInvoice.receipt_no,
        'flag'         : 'invoice' if type == 'invoice' else 'receipt',
        'invoiceObj'   : getInvoice,
        'senangpay_url': settings.PROD_SENANGPAY_URL + settings.PROD_SENANGPAY_MERCHANT_KEY,
        'invoice_no'   : getInvoice.invoice_no.replace("/", "-"),
        'host_address' : settings.HOST_ADDRESS,
        'companySalt'  : GenericLibraries().saltEncode(getCompanyObj)
    }
    return render(request, 'InvoiceAndPayment/InvoiceView.html', context)

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def CancelInvoiceView(request):
    try:
        getInvoice      = Invoice.objects.get(invoice_no=request.POST['invoiceNo'])
        getOrderObj     = Order.objects.filter(invoice_id=getInvoice.id)
        for order in getOrderObj:
            member = Member.objects.get(id=order.member_id)
            getAdPremiumObj = PremiumAdjustment.objects.filter(member_id = member.id, paid = False, generated_invoice = True, void = False)
            if getAdPremiumObj:
                for adPremium in getAdPremiumObj:
                    adPremium.generated_invoice = False
                    adPremium.save()
            if member.renew:
                member.status = 'Pending Payment'
            elif member.paid:
                member.status = 'Active'
            else:
                member.status = 'Accept'
            member.generated_invoice = False
            member.save()
            
        getInvoice.status  = "Void"
        getInvoice.remarks = request.POST['cancelInvoiceReason']
        getInvoice.updated_datetime = str(datetime.datetime.now())
        getInvoice.save()
    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/invoice-list')
    return redirect(settings.HOST_ADDRESS+'/invoice-list')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def InvoiceListView(request):
    try:
        getCompanyObj   = CorporateProfile.objects.get(user_id=request.user.id)
        getInvoice      = Invoice.objects.filter(company_id=getCompanyObj.id)
        invoiceDict     = []
        for inv in getInvoice:
            serializerINV = json.loads(serializers.serialize('json', [inv]))
            serializerINV[0]['fields'].update({
                'id'            : inv.id,
                'invoice_no'    : inv.invoice_no,
                'invoiceNo'     : inv.invoice_no.replace("/", "-"),
                'ref_no'        : inv.senangpay_refno,
                'amount'        : inv.total_amount,
                'status'        : inv.status,
                'hash_value'    : inv.hash_value,
                'invoiceSalt'   : GenericLibraries().saltEncode(inv)
            })
            if inv.updated_datetime:
                serializerINV[0]['fields'].update({
                'cancellation_date' : inv.updated_datetime
            })
            invoiceDict.append(serializerINV[0]['fields'])
        invoiceTable = Paginator(invoiceDict, 10)

        context = {
            'invoices' : invoiceTable,
            'host_address' : settings.HOST_ADDRESS,
            'senangpay_url': settings.PROD_SENANGPAY_URL + settings.PROD_SENANGPAY_MERCHANT_KEY,
            'company'   : getCompanyObj,
            'companySalt'  : GenericLibraries().saltEncode(getCompanyObj)
        }

        if request.method == 'POST':
            filter_keys = {}
            searchDate = ''
            searchStatus = ''
            for k, v in request.POST.items():
                if k.startswith("searchInvoice"):
                    if k.removeprefix("searchInvoice") == "Date":
                        searchDate = v
                        context['filterDate'] = v
                    elif k.removeprefix("searchInvoice") == "Status":
                        searchStatus = v
                        context['filterStatus'] = v
            if searchDate:
                searchDate = searchDate.replace("-", "")
                filter_keys['invoice_no__icontains'] = "N"+searchDate
            if searchStatus:
                filter_keys['status'] = searchStatus

            getFilteredInvoice = Invoice.objects.filter(**filter_keys, company_id=getCompanyObj.id)
            filteredInvoiceDict = []
            for inv2 in getFilteredInvoice:
                serializerINV = json.loads(serializers.serialize('json', [inv2]))
                serializerINV[0]['fields'].update({
                    'id'            : inv2.id,
                    'invoice_no'    : inv2.invoice_no,
                    'ref_no'        : inv2.senangpay_refno,
                    'amount'        : inv2.total_amount,
                    'status'        : inv2.status,
                    'hash_value'    : inv2.hash_value,
                    'invoiceSalt'   : GenericLibraries().saltEncode(inv2)
                })
                filteredInvoiceDict.append(serializerINV[0]['fields'])
            filteredInvoiceTable = Paginator(filteredInvoiceDict, 10)
            context['invoices'] = filteredInvoiceTable
        return render(request, 'InvoiceAndPayment/InvoiceList.html', context)
    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/member-list')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def ExportCorporatelist(request):
    try:
        corporate_list = CorporateProfile.objects.filter(rejected=False)
        # Create an Excel workbook and add a worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Add headers to the worksheet
        ws.append(['#', 'Company Name', 'Company Register Number', 'Company Contact Number', 'Company Email', 'Company Address', 'Payment Due Date', 'Status', 'Payment Mode', 'Relationship'])  

        auto_increment = 1
        for corporate in corporate_list:
            address_parts = [part for part in [corporate.address_line1, corporate.address_line2, corporate.address_line3, corporate.state, corporate.city, corporate.postcode] if part is not None]
            address = " ".join(address_parts)

            # Access the related CompanyRelationship instance
            try:
                company_relationship = CompanyRelationship.objects.filter(company=corporate).first()
                if company_relationship:
                    relationship_name = company_relationship.get_relationship_name()
                else:
                    relationship_name = "-"
            except CompanyRelationship.DoesNotExist:
                relationship_name = "-"

            row_data = [
            auto_increment,
            corporate.company_name if corporate.company_name else "-",
            corporate.registration_no if corporate.registration_no else "-",
            corporate.contact1 if corporate.contact1 else "-",
            corporate.email_address if corporate.email_address else "-",
            address if address else "-",
            corporate.payment_due_date if corporate.payment_due_date else "-",
            corporate.status if corporate.status else "-",
            corporate.payment_mode if corporate.payment_mode else "-",
            relationship_name  
            ]
            ws.append(row_data)
            auto_increment += 1 

        # Auto-fit column height
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column) 
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add padding
            ws.column_dimensions[column].width = adjusted_width

        # Create an HttpResponse to serve the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Corporate List.xlsx"'

        wb.save(response)
        return response
    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/company-approval')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def ExportRejectedCorporatelist(request):
    try:
        corporate_list = CorporateProfile.objects.filter(rejected=True)
        # Create an Excel workbook and add a worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Add headers to the worksheet
        ws.append(['#', 'Rejected Reason', 'Company Name', 'Company Register Number', 'Company Contact Number', 'Company Email', 'Company Address', 'Payment Due Date', 'Status', 'Payment Mode', 'Relationship'])  

        auto_increment = 1
        for corporate in corporate_list:
            address_parts = [part for part in [corporate.address_line1, corporate.address_line2, corporate.address_line3, corporate.state, corporate.city, corporate.postcode] if part is not None]
            address = " ".join(address_parts)

            # Access the related CompanyRelationship instance
            try:
                company_relationship = CompanyRelationship.objects.filter(company=corporate).first()
                if company_relationship:
                    relationship_name = company_relationship.get_relationship_name()
                else:
                    relationship_name = "-"
            except CompanyRelationship.DoesNotExist:
                relationship_name = "-"

            row_data = [
            auto_increment,
            corporate.remarks if corporate.remarks else "-",
            corporate.company_name if corporate.company_name else "-",
            corporate.registration_no if corporate.registration_no else "-",
            corporate.contact1 if corporate.contact1 else "-",
            corporate.email_address if corporate.email_address else "-",
            address if address else "-",
            corporate.payment_due_date if corporate.payment_due_date else "-",
            corporate.status if corporate.status else "-",
            corporate.payment_mode if corporate.payment_mode else "-",
            relationship_name  
            ]
            ws.append(row_data)

            auto_increment += 1 

        # Auto-fit column height
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column) 
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add padding
            ws.column_dimensions[column].width = adjusted_width

        # Create an HttpResponse to serve the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Rejected Corporate List.xlsx"'

        wb.save(response)
        return response
    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/company-approval')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def ExportMemberList(request):
    try:
        if request.method == 'POST':
            company_id = request.POST.get('company_id')
            corporate = CorporateProfile.objects.get(id=company_id)
            members  = Member.objects.filter(corporate_id=company_id)

            # Create an Excel workbook and add a worksheet
            wb = openpyxl.Workbook()
            ws = wb.active

            # Add headers to the worksheet
            ws.append(['#', 'Member Name', 'Member Email', 'Member Contact Number', 'Member Nationality', 'Member DOB', 'Member Gender', 'Member MyKad', 'Member Batch No', 'Status', 'Death Coverage', 'Disability Coverage', 'Critical Illness Coverage', 'Accident Coverage', 'Medical Coverage', 'Premium (RM)', 'Invoice Date', 'Payment Due Date'])  

            product_ids = [1, 2, 3, 4, 5]
            coverages = {product_id: {} for product_id in product_ids}

            for product_id in product_ids:
                product_coverages = MemberProductMapping.objects.filter(product=product_id, member__in=members)
                for member in members:
                    matching_coverage = product_coverages.filter(member=member).first()
                    if matching_coverage:
                        coverages[product_id][member.id] = matching_coverage.coverage_amount
                    else:
                        coverages[product_id][member.id] = '-'

            auto_increment = 1
            for member in members: 
                try:
                    order = Order.objects.filter(member_id=member.id).order_by('-created_datetime').first()

                    if order:
                        invoice = Invoice.objects.exclude(status='Void').get(id=order.invoice_id)
                        invoice_date = invoice.created_datetime.strftime('%Y-%m-%d')
                    else:
                        invoice_date = '-'

                except Invoice.DoesNotExist:
                    # Handle the case when the invoice with the provided ID does not exist or is "Void"
                    invoice_date = '-'

                row_data = [
                auto_increment,
                member.name,
                member.email_address,
                member.mobile_no,
                member.nationality,
                member.dob,
                member.gender,
                member.mykad,
                member.batch_no,
                member.status,
                coverages[1][member.id],  # Death Coverage
                coverages[2][member.id],  # Disability Coverage
                coverages[3][member.id],  # Critical Illness Coverage
                coverages[4][member.id],  # Accident Coverage
                coverages[5][member.id],  # Medical Coverage
                member.quotation_premium,
                invoice_date,
                corporate.payment_due_date
                ]
                ws.append(row_data)

                auto_increment += 1 

            for col in ws.columns:
                for cell in col:
                    cell.alignment = Alignment(horizontal='left')
            #align amount column to right      
            for col in ws.iter_cols(min_row=2, min_col=11, max_col=16):  # Columns K to P, starting from the second row
                for cell in col:
                    cell.alignment = Alignment(horizontal='right')
            # Auto-fit column height
            for col in ws.columns:
                max_length = 0
                column = get_column_letter(col[0].column) 
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 3)  
                ws.column_dimensions[column].width = adjusted_width
            ws.column_dimensions['Q'].width = max_length + 5  # value too long for invoice date

            # Create an HttpResponse to serve the Excel file
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Member List.xlsx"'

            wb.save(response)
            return response
        context = {
            'company': company_id,
            'members': members,
        }

        return render(request, 'CorporateApproval/MemberList.html', context)
    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/company-approval')
    
@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def DownloadTemplates(request, template_name):
    try:
        previous_url = resolve(request.GET['path_name']).url_name
        with open(settings.STATICFILES_DIRS[0] + '/download/' + template_name, 'rb') as f:
            data = f.read()

        response                        = HttpResponse(data, content_type='text/xlsx')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(template_name)

        return response
    except IOError:
        messages.error(request, 'Template not found!')
        return redirect(settings.HOST_ADDRESS + '/' + previous_url)

def DownloadForm(request, form_name):
    try:
        previous_url = resolve(request.GET['path_name']).url_name
        with open(settings.STATICFILES_DIRS[0] + '/download/' + form_name, 'rb') as f:
            data = f.read()
        
        response                        = HttpResponse(data, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(form_name)

        return response
    except IOError:
        messages.error(request, 'Template not found!')
        return redirect(settings.HOST_ADDRESS + '/' + previous_url)

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def DownloadFailedUploads(request, fileName, companySalt):
    try:
        decodeCompanySalt = base64.b64decode(companySalt)
        decodeASCII = decodeCompanySalt.decode('UTF-8')
        getCompanyID = decodeASCII.split('_')
        companyID = getCompanyID[1]
        filePath = settings.MEDIA_ROOT + "/" + CorporateProfile.objects.get(user_id=request.user.id).company_name + "/FailUploadMember/" + fileName
        with open(filePath, 'rb') as f:
            data = f.read()

        response                        = HttpResponse(data, content_type='text/xlsx')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(fileName)

        return response

    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/failed-upload-list/{}'.format(companySalt))

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def DownloadPreviousUploads(request, fileName, companySalt):
    try:
        decodeCompanySalt = base64.b64decode(companySalt)
        decodeASCII = decodeCompanySalt.decode('UTF-8')
        getCompanyID = decodeASCII.split('_')
        companyID = getCompanyID[1]
        filePath = settings.MEDIA_ROOT + "/" + CorporateProfile.objects.get(user_id=request.user.id).company_name + "/Upload/" + fileName
        with open(filePath, 'rb') as f:
            data = f.read()

        response                        = HttpResponse(data, content_type='text/xlsx')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(fileName)

        return response

    except Exception as e:
        messages.error(request, str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/failed-upload-list/{}'.format(companySalt))


def convert_excel_to_pdf(excel_file):
    pdf_buffer = BytesIO()
    workbook = openpyxl.load_workbook(excel_file)
    ws = workbook.active
    
    for img in ws._images:
        img_data = img.image
        with BytesIO(img_data) as image_buffer:
            image = Image(image_buffer)
            img.image = image
    
    workbook.save(pdf_buffer)
    return pdf_buffer  

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS,redirect_field_name=None)
def ViewUploadedFile(request,fileName,companySalt):
    try:
        decodeCompanySalt = base64.b64decode(companySalt)
        decodeASCII = decodeCompanySalt.decode('UTF-8')
        getCompanyID = decodeASCII.split('_')
        getCompanyObj       = CorporateProfile.objects.get(user_id=request.user.id)
        # filePath = settings.MEDIA_ROOT + "/" + CorporateProfile.objects.get(user_id=request.user.id).company_name + "/Upload/" + fileName
        # pdfFileName = os.path.splitext(fileName)[0]+'.pdf'
        filePath = os.path.join(settings.MEDIA_ROOT, getCompanyObj.company_name, "Upload", fileName)
        # os.chmod(filePath,S_IREAD)
        
        # Read the contents of the excel file using openpyxl
        # excel_data = []
        # workbook = openpyxl.load_workbook(filePath, read_only=True)
        # for sheets in workbook.sheetnames:
        #     worksheet = workbook[sheets]
        #     sheet_data = []
        #     skip_row =  {1,4,5}
        #     header_row = None
        #     for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start = 2):
        #         if not header_row:
        #             header_row = row
        #         elif not skip_row.intersection({row_index}):
        #             if all(cell is None for cell in row):
        #                 break
        #             else:
        #                 sheet_data.append(row)
        
        #     excel_data.append({'sheet_name': sheets, 'header-row': header_row, 'sheet_data': sheet_data})
            
        pdf_data = convert_excel_to_pdf(filePath)
        
        pdf_filename = os.path.splitext(fileName)[0]+'.pdf'
        pdf_file_path = os.path.join(settings.MEDIA_ROOT, pdf_filename)
        with open(pdf_file_path, 'wb') as pdf_file:
            pdf_file.write(pdf_data.getbuffer())
            
        pdf_url = settings.MEDIA_URL + pdf_filename
        
        # Win32com Client
        # excel = client.Dispatch('Excel.Application',pythoncom.CoInitialize())
        # sheets = excel.Workbooks.Open(filePath)
        # work_sheets = sheets.Worksheets[0]
        # work_sheets.PageSetup.Orientation = 2
        # work_sheets.PageSetup.Zoom = False
        # work_sheets.PageSetup.FitToPagesWide = 1

        # AsposeCells
        # workbook = Workbook(filePath)
        # pdfOptions = PdfSaveOptions()
        # pdfOptions.setCompliance(PdfCompliance.PDF_A_1_A)
        # workbook.save(settings.MEDIA_ROOT + "/" + CorporateProfile.objects.get(user_id=request.user.id).company_name + "/Upload/" +pdfFileName,pdfOptions)

        # Pdfkit
        # path_wkthmltopdf = 'C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe'
        # config = pdfkit.configuration(wkhtmltopdf=path_wkthmltopdf)
        # df = pd.read_excel(filePath, index_col= 0, skiprows=1)
        # df.to_html('file.html')
        # pdfkit.from_file('file.html',settings.MEDIA_ROOT + "/" + CorporateProfile.objects.get(user_id=request.user.id).company_name + "/Upload/" +pdfFileName,configuration=config)
        
        # run libreoffice in ubuntu
        # cmd = 'libreoffice --headless --convert-to pdf '+filePath+' --outdir '+settings.MEDIA_ROOT + "/" + CorporateProfile.objects.get(user_id=request.user.id).company_name + "/Upload/"
        # os.system(cmd)

        # work_sheets.ExportAsFixedFormat(0,settings.MEDIA_ROOT + "/" + CorporateProfile.objects.get(user_id=request.user.id).company_name + "/Upload/" + pdfFileName)

        context={
            'company'    : getCompanyObj,
            'fileName'   : fileName,
            # 'pdfFile'    : pdfFileName,
            'companySalt': companySalt,
            # 'excel_data' : excel_data, #parsed Excel data
            'pdf_url' : pdf_url,
        }
        
        return render(request,'Menu/ViewUploadedFile.html',context)
    
    except Exception as e:
        messages.error(request,str(e))
        logger.error(str(e),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/upload-history/{}'.format(companySalt))

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None) 
def CompanyModifyPDDListView(request):
    try:
        rowCount = 5
        getVerifiedCompany = CorporateProfile.objects.filter(verified=True)
        verifiedCompanyLists = []
        getCorpObj = GetCompanies()

        for cmp in getVerifiedCompany:
            serializerCMP    = json.loads(serializers.serialize('json', [cmp]))
            
            if not cmp.rejected:
                company_details = serializerCMP[0]['fields']

                compExistCurrentDate = CurrentDate.objects.filter(corporate_id=cmp.id).first()
                
                if compExistCurrentDate:
                    # Append the existing current date to the company details
                    company_details['current_date'] = datetime.datetime.strftime(compExistCurrentDate.current_datetime, '%Y-%m-%d')
                else:
                    # Use the current system date and append it to the company details
                    company_details['current_date'] = datetime.date.today().strftime('%Y-%m-%d')
                company_details['company_id'] = cmp.id
                verifiedCompanyLists.append(company_details)

        companyTable = Paginator(verifiedCompanyLists, rowCount)

        context = {
            'company_lists'          : companyTable,
            'path'                   : request._current_scheme_host,
            'host_address' : settings.HOST_ADDRESS
        }
        
        if request.method == 'POST':
            for k, v in request.POST.lists():
                if 'searchCompany' in k:
                    print(request.POST['searchCompany'])
                    context['company_lists'] = getCorpObj.getCompanies(company_name__icontains=request.POST['searchCompany'])
                    context['filterCompany'] = request.POST['searchCompany']
        return render(request, 'CorporateChangePaymentDD/UpdateCorporatePDD.html', context)
        
    except Exception as ex:
        messages.error(request, str(ex))
        logger.error(str(ex),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/company-approval')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS,redirect_field_name=None)
def CompanyAmmendPDD(request):
    try:
        if request.method == 'POST':
            deartimeUserID = request.POST.get('deartimeUserID')
            deartimePayerID = request.POST.get('deartimePayerID')
            ammendDayPDD = request.POST.get('ammendDayPDD')
            ammendMonthPDD = request.POST.get('ammendMonthPDD')
            current_year = datetime.datetime.now().year
            new_due_date = f"{current_year}-{ammendMonthPDD}-{ammendDayPDD}"
            new_due_date = datetime.datetime.strptime(new_due_date, '%Y-%m-%d')
            deartimeDB  = DearTimeDbConn()
            isConnected = deartimeDB.connect()
            if not isConnected:
                messages.error(request, settings.CONNECTION_LOST_MESSAGE)
                return redirect(settings.HOST_ADDRESS+'/company-modify-PDD')
            else:
                corporateCoveragesNDD = deartimeDB.exec_SQL('getCoverageNDD', {'PAYER_ID':deartimePayerID }, 'fetchone')
                lastPaymentOn = corporateCoveragesNDD['dset'][corporateCoveragesNDD['dcolname'].index('last_payment_on')]
                if lastPaymentOn > new_due_date:
                    new_due_date = new_due_date + relativedelta(years=1)
                updateCompanyNDD = {'UPDATED_DATE':str(datetime.datetime.now()), 'NDD_PAYMENT_DUE_DATE': new_due_date, 'PAYER_ID':deartimePayerID}
                deartimeDB.exec_SQL('updateCoveragesNDD', updateCompanyNDD, 'update')    
                getCompanyObj = CorporateProfile.objects.get(user=deartimeUserID)
                getCompanyObj.payment_due_date = datetime.datetime.strftime(new_due_date, '%Y-%m-%d')
                getCompanyObj.save()
                messages.success(request, 'Payment due date updated successfully.')
                return redirect(settings.HOST_ADDRESS+'/company-modify-PDD')
        else:
            return redirect(settings.HOST_ADDRESS+'/company-modify-PDD')
    except Exception as ex:
        messages.error(request, str(ex))
        logger.error(str(ex),extra={'username':request.user.id})
        return redirect(settings.HOST_ADDRESS+'/company-approval')
    
@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS,redirect_field_name=None)
def CompanyAmmendCD(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'save':
            corporateID = request.POST.get('corporate_id')
            # ammendDayCD = request.POST.get('cdayDropdown')
            # ammendMonthCD = request.POST.get('cmonthDropdown')
            # currentYear = datetime.datetime.now().year
            # # Create a new date object with the amended day and month
            # new_date = f"{currentYear}-{ammendMonthCD}-{ammendDayCD}"
            # new_date = datetime.datetime.strptime(new_date, '%Y-%m-%d')
            new_current_date = request.POST.get('currentDate')

            try:
                currentDate = CurrentDate.objects.get(corporate_id=corporateID)
                currentDate.current_datetime = new_current_date
                currentDate.updated_datetime = datetime.datetime.now()
                currentDate.save()
            except:
                # Create a new CurrentDate record if it doesn't exist
                CurrentDate.objects.create(corporate_id=corporateID, current_datetime=new_current_date)
            return redirect(settings.HOST_ADDRESS + '/company-modify-PDD') 
        # elif action == 'reset':
        #     corporateID = request.POST.get('corporate_id')
        #     try:
        #         current_date = CurrentDate.objects.get(corporate_id=corporateID)
        #         current_date.delete()
                
        #         return redirect(settings.HOST_ADDRESS + '/company-modify-PDD') 
        #     except CurrentDate.DoesNotExist:
        #         # Handle the case where the data does not exist
        #         return redirect(settings.HOST_ADDRESS + '/company-modify-PDD') 

        return redirect(settings.HOST_ADDRESS + '/company-approval')            

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS,redirect_field_name=None)
def CompanyResetCD(request):
    if request.method == 'POST':
        corporateID = request.POST.get('corporate-id')
        try:
            current_date = CurrentDate.objects.get(corporate_id=corporateID)
            current_date.delete()
            
            return redirect(settings.HOST_ADDRESS + '/company-modify-PDD') 
        except CurrentDate.DoesNotExist:
            # Handle the case where the data does not exist
            return redirect(settings.HOST_ADDRESS + '/company-modify-PDD') 
    return redirect(settings.HOST_ADDRESS + '/company-approval')       

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS,redirect_field_name=None)
def ClearUsers(request):
    if request.method == 'POST':
        userID_list = request.POST.get('memberIdField', '').split(',')
        userID_list = [userID for userID in userID_list if userID]
        
        if not userID_list:
            messages.error(request, 'No users selected for clearing!')
            return render(request, 'ClearRecords/DeleteBase.html')

        try:
            with transaction.atomic():
                for count, userID in enumerate(userID_list, start=1):
                    user = Member.objects.get(id=userID)
                    random_code = GenericLibraries().code_randomiser()
                    user.email_address += random_code + str(count)
                    user.mobile_no += random_code + str(count)
                    user.mykad += random_code + str(count)
                    user.save()

            messages.success(request, 'Users cleared successfully!')
            return render(request, 'ClearRecords/DeleteBase.html')

        except Exception as e:
            messages.error(request, str(e))
            logger.error(str(e), extra={'username': request.user.id})
            return redirect(settings.HOST_ADDRESS)

    else:
        messages.error(request, 'Invalid request method!')
        return render(request, 'ClearRecords/DeleteBase.html')

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS,redirect_field_name=None)
def DeleteUser(request):
    try:
        if request.method == 'POST':
            searchBy = request.POST.get('searchBy')
            userInfo = request.POST.get('search-user')
            if searchBy == 'mykad':
                userFound = Member.objects.filter(mykad=userInfo)
            elif searchBy == 'email':
                userFound = Member.objects.filter(email_address=userInfo)
            elif searchBy == 'phone':
                userFound = Member.objects.filter(mobile_no=userInfo)
            if not userFound:
                messages.error(request, 'No records found.')
            else:
                userFound = userFound.select_related('corporate')
            return render(request, 'ClearRecords/DeleteBase.html', {'userFound': userFound})
        else:
            return render(request, 'ClearRecords/DeleteBase.html')
    except Exception as e:
        messages.error(request,str(e))
        logger.error(str(e),extra={'username':request.user.id})        

@csrf_exempt
@login_required(login_url=settings.HOST_ADDRESS, redirect_field_name=None)
def LogOutView(request):
    logout(request)
    return redirect(settings.HOST_ADDRESS)
